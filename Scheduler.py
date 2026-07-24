# =======================================================
# Program Status
# Last Updated: 07-13-2026 12:41 PM
# Author: Logan Burkardt
#
# Current Functionality:
# - Reads Open Order Report (OOR).xlsx
# - Identifies jobs ready for molding
# - Excludes:
#     * On Hold jobs
#     * Already Scheduled jobs
#     * Jobs with 0 molds needed
#     * Investment castings (IFA, IFC, I)
#
# Planned Development:
# - Push mold schedule into daily production schedule
# - Build Melt WIP schedule
# - Build Casting/Cleaning WIP schedule
# - Automate database exports
# - Consolidate file repositories into a single location
# =======================================================


# =======================================================
# Description:
# This script reads the Open Order Report (OOR), filters jobs ready for molding,
# and builds a mold schedule while considering daily mold limits.
# It also provides functions to expand jobs into multiple extensions,
# assign schedule days, and print the schedule by bucket.
# =======================================================

# =======================================================
# Concerns and considerations for developement
# 1. How to ensure if an order is partially poured it is still scheduled until complete. 
# and builds a mold schedule while considering daily mold limits.
# It also provides functions to expand jobs into multiple extensions,
# assign schedule days, and print the schedule by bucket.
# =======================================================


# =======================================================
# Need to Refactor
# scheduler/
# │
# ├── Scheduler.py          # Main entry point only
# ├── config.py             # Constants and settings
# ├── schedule_logic.py     # Filtering, splitting, day assignment
# ├── schedule_builder.py   # Daily schedule creation
# ├── exports.py            # Excel export functions
# ├── io_utils.py           # File reading
# └── models.py             # Optional later


# =======================================================
# Imports
# =======================================================

import math
import string
from datetime import datetime, timedelta

import pandas as pd

# May be removed if source data remains consistent
# from fuzzywuzzy import process  

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# Saving for later - need to create a GUI for users for easier interaction with the scheduler.
# import PyQT6 

#consider moving from config to constants?
from config import Columns
from config import DailyMoldLimits

# ========================================================
# MOVED GLOBAL VARIABLES TO CONFIG.PY
# ========================================================

# dictionary to track job counts as the filtering
# incremented during mold scheduler
filtered_job_counts = {
    "blank":0,
    "hold":0,
    "scheduled":0,
    "job_type":0,
    "cast_type":0,
    "no_molds":0,
    "added":0,
}

schedule_days = {
    1: {"L": 0, "F": 0},
    2: {"L": 0, "F": 0}
}

EXPORT_COLUMNS = {
    Columns.COL_DUE_DATE: "Due Date",
    "Customer Name": "Customer Name",
    "Part Number": "Part Number",
    Columns.COL_JOB_NUMBER: "Job Number",
    "EXT": "EXT",
    Columns.COL_ALLOY: "Alloy",
    Columns.COL_CAST_TYPE: "Mold Type",
    "Quantity of Molds": "Quantity of Molds",
    "Castings Per Mold": "Castings Per Mold",
    "Quantity of Cores": "Cores Per Mold",
    "Total Weight per EXT": "Total Weight per EXT",
    "Molds for EXT": "# of Molds for EXT"
}

# // =======================================================
# // Functions
# // =======================================================

def Read_File(filepath = "C:\\Users\\lburkardt\\OneDrive - MonettMetalsUS1\\Quality\\Schedule\\Open Order Report.xlsx"):
    # // Read input from file for list of ready to schedule jobs
    # File 
    ImportedFile = pd.read_excel(filepath, sheet_name="OOR")

    # remove trailing and leading white space
    ImportedFile.columns = ImportedFile.columns.str.strip()

    #test print
    # print(ImportedFile)  

    return(ImportedFile)


def Mold_Scheduler(ReadyToMold):
    # Molds needed column
    # =====================================================================
    # Filters the Open Order Report down to jobs that are eligible
    # for mold scheduling.
    #
    # Removes:
    #     - Blank rows
    #     - On Hold jobs
    #     - Already scheduled jobs
    #     - Investment casting jobs
    #     - Jobs requiring no molds
    # 
    # Returns:
    #     List of schedulable job rows.
    #
    # =====================================================================

    jobs_to_schedule = []

    for row, job in ReadyToMold.iterrows():

        # Filter out blank rows
        if pd.isna(job[Columns.COL_JOB_NUMBER]): 
            filtered_job_counts["blank"] += 1
            continue
         
        # filter checks for On hold, already scheduled, or all molds completed, or if investment job by job type or cast type
        if str(job[Columns.COL_HOLD]).upper() == "YES":
            filtered_job_counts["hold"] += 1
            continue
        
        # Filters out investments jobs based on job type
        if str(job[Columns.COL_JOB_TYPE]).upper() in ["IFA", "IFC"]:
            filtered_job_counts["job_type"] += 1
            continue
        
        # filters out jobs that are already scheduled
        if str(job[Columns.COL_SCHEDULED]).upper() == "YES":
            filtered_job_counts["scheduled"] += 1
            continue

        # filters out investment casting jobs based on cast type
        if str(job[Columns.COL_CAST_TYPE]).upper() == "I":
            filtered_job_counts["cast_type"] += 1
            continue
            
        # filters out jobs that require no molds
        if job[Columns.COL_MOLDS_NEEDED] <= 0:
            filtered_job_counts["no_molds"] += 1
            continue
        
        # add jobs that made it through the filter to jobs to schedule list
        filtered_job_counts["added"] += 1
        jobs_to_schedule.append(job)


    return jobs_to_schedule
        
# function to assign extension letter based on the number of splits required for a job  
def get_extensions(num_splits):
    if num_splits == 1:
         return [""]
    
    # create list to assign extensions for that specific job - will need called for each row?
    extensions = []

    # if only one split is needed, return "L" as the extension
    alphabet = list(string.ascii_uppercase)

    # Assign extensions for each split of the job
    for i in range(num_splits - 1):
        extensions.append(alphabet[i])

    # Append "L" as the last extension for the job
    extensions.append("L")

    return extensions

# calculate how many times we need to split a job (basically how many days to complete a job so we can assign extension letters)
def Calculate_Splits(job):
    
    # calculate the number of splits required for the job based on molds needed and daily limit
    molds_needed = math.ceil(job[Columns.COL_MOLDS_NEEDED])
    
    # round up mold count as we cannot produce partial molds
    daily_limit = Get_daily_mold_limit(job)
    
    splits = math.ceil(molds_needed / daily_limit)

    return splits

# prep job for push into daily mold schedule format
def Expand_Job(job):
    # prepare job for expansion into multiple schedule rows based on splits and extensions
    molds_needed = math.ceil(job[Columns.COL_MOLDS_NEEDED])

    splits = Calculate_Splits(job)

    extensions = get_extensions(splits)

    # initialize list to hold expanded rows for each extension
    rows = []

    # track the number of molds remaining to be scheduled for this job
    molds_remaining = molds_needed

    # determine the daily mold limit for this job
    daily_limit = Get_daily_mold_limit(job)

    # find which is lesser, molds needed or max molds
    for seq, ext in enumerate(extensions):

        molds_for_ext = min(
            daily_limit, molds_remaining
        )
        # create a new row for this extension with the calculated molds and other details
        row = job.copy()

        row["EXT"] = ext
        row["Extension_Seq"] = seq
        row["Molds for EXT"] = molds_for_ext

    
        row["Total Weight per EXT"] = (
            molds_for_ext *
            row[Columns.COL_POUR_WEIGHT]
        )

        # add the expanded row to the list of rows for this job extension
        rows.append(row)

        molds_remaining -= molds_for_ext

    return rows

# add jobs to list after splitting into extensions
def Build_Schedule_Rows(jobs_to_schedule):
    schedule_rows = []

    for job in jobs_to_schedule:
        expanded_rows = Expand_Job(job)

        schedule_rows.extend(expanded_rows)

    return schedule_rows

# test function to print out selected jobs to schedule to mold
def jobs_to_schedule_test():
    print("\nJobs selected for scheduling:")

    for job in Jobs_to_schedule:
        print(
            f"{job[config.COL_JOB_NUMBER]} | "
            f"{job['Customer Name']} | "
            f"Molds Needed: {job[config.COL_MOLDS_NEEDED]}"
        )

    print(f"\nTotal Jobs Selected: {len(Jobs_to_schedule)}")

# test function to print out selected jobs with extensions assigned.
def Scheduled_rows_test():
    
    for row in Schedule_rows:
        print(
            f"{row[config.COL_JOB_NUMBER]}"
            f"{row['EXT']} | "
            f"{row['Molds for EXT']} molds"
        )


# =====================================================================================================
# CAUTION BELOW HERE IS UNDOCUMENTED AI SLOP, BUT IT WORKS. WILL REFACTOR AND COMMENT WHEN TIME ALLOWS.
# =====================================================================================================

# checks row for casting type and assigns limit per day that can be made
def Get_daily_mold_limit(job):
    pour_weight = job[Columns.COL_POUR_WEIGHT]

    if pour_weight > 300:
        return 3
    
    casting_type = str(job[Columns.COL_CAST_TYPE]).upper()

    if casting_type == "F":
        return 3
    
    return 6


def Is_F_Job(job):

    if job[Columns.COL_POUR_WEIGHT] > 300:
        return True

    return str(job[Columns.COL_CAST_TYPE]).upper() == "F"


def Assign_days(schedule_df):

    schedule_df["Schedule Day"] = None

    day_usage = {}

    job_last_day = {}

    part_usage = {}


    for index, row in schedule_df.iterrows():

        molds = row["Molds for EXT"]

        bucket = "F" if Is_F_Job(row) else "L"

        job_num = row[Columns.COL_JOB_NUMBER]

        part_num = row["Part Number"]

        day = job_last_day.get(job_num, 0) + 1

        while True:
            
            if day not in day_usage:
                day_usage[day] = {
                    "L": 0,
                    "F": 0
                }

            if day not in part_usage:
                part_usage[day] = {}

            if part_num not in part_usage[day]:
                part_usage[day][part_num] = 0
            
            capacity = (
                DailyMoldLimits.MAX_F_MOLDS_PER_DAY
                if bucket == "F"
                else DailyMoldLimits.MAX_L_MOLDS_PER_DAY
            )

            if (
                day_usage[day][bucket] + molds <= capacity
                and
                part_usage[day][part_num] + molds <= 6
                ):

                day_usage[day][bucket] += molds

                part_usage[day][part_num] += molds

                schedule_df.at[index, "Schedule Day"] = day

                job_last_day[job_num] = day

                break

            day += 1

    print(day_usage)

    return schedule_df


def print_bucket(Schedule_Data_Frame):

    for day in sorted(
        Schedule_Data_Frame["Schedule Day"].unique()
    ):

        day_rows = Schedule_Data_Frame[
            Schedule_Data_Frame["Schedule Day"] == day
        ]

        l_molds = day_rows[
            ~day_rows.apply(Is_F_Job, axis=1)
        ]["Molds for EXT"].sum()

        f_molds = day_rows[
            day_rows.apply(Is_F_Job, axis=1)
        ]["Molds for EXT"].sum()

        print(
            f"Day {day}: "
            f"L={l_molds}/{DailyMoldLimits.MAX_L_MOLDS_PER_DAY}, "
            f"F={f_molds}/{DailyMoldLimits.MAX_F_MOLDS_PER_DAY}"
            f"F={f_molds}/{config.max_f_molds_per_day}"
        )

def Build_Daily_Schedules(Schedule_Data_Frame):

    daily_schedules = {}

    for day in sorted(Schedule_Data_Frame["Schedule Day"].unique()):
        daily_schedules[day] = (
            Schedule_Data_Frame[Schedule_Data_Frame["Schedule Day"] == day]
            .copy()
            .sort_values(
                by=[Columns.COL_ALLOY,Columns.COL_JOB_NUMBER]
            )
        )
    return daily_schedules

def Build_Daily_Export_Blocks(
    Daily_Schedules,
    Day_Dates
):

    export_blocks = {}

    for day, df in Daily_Schedules.items():

        weight_total = (
            df["Total Weight per EXT"]
            .fillna(0)
            .sum()
        )

        mold_total = (
            df["Molds for EXT"]
            .fillna(0)
            .sum()
        )

        export_blocks[day] = {

            "date": Day_Dates[day]["date"],
            "weekday": Day_Dates[day]["weekday"],

            "rows": df[
                [
                    Columns.COL_DUE_DATE,
                    "Customer Name",
                    "Part Number",
                    Columns.COL_JOB_NUMBER,
                    "EXT",
                    Columns.COL_ALLOY,
                    Columns.COL_CAST_TYPE,
                    "Quantity of Molds",
                    "Castings Per Mold",
                    "Quantity of Cores",
                    "Total Weight per EXT",
                    "Molds for EXT"
                ]
            ].copy(),

            "weight_total": weight_total,
            "mold_total": mold_total
        }

    return export_blocks

def Build_Schedule_Dates(
    daily_schedules,
    start_date
):

    day_dates = {}

    current_date = start_date

    for day in sorted(daily_schedules.keys()):

        while current_date.weekday() > 4:
            current_date += timedelta(days=1)

        day_dates[day] = {
            "date": current_date,
            "weekday": current_date.strftime("%A")
        }

        current_date += timedelta(days=1)

    return day_dates

def Print_Export_Blocks(export_blocks):

    for day in export_blocks:

        print("\n" + "=" * 50)

        print(
            f"Mold Schedule    "
            f"{export_blocks[day]['date'].strftime('%m/%d/%Y')}    "
            f"{export_blocks[day]['weekday']}"
        )

        print("=" * 50)

        print(
            export_blocks[day]["rows"]
        )

        print(
            f"\nTOTAL WEIGHT: "
            f"{export_blocks[day]['weight_total']}"
        )

        print(
            f"TOTAL MOLDS: "
            f"{export_blocks[day]['mold_total']}"
        )

def Build_Excel_Rows(export_blocks):

    excel_rows = []

    for day in sorted(export_blocks.keys()):

        block = export_blocks[day]

        excel_rows.append([
            "Mold Schedule",
            block["date"].strftime("%m/%d/%Y"),
            block["weekday"]
        ])

        excel_rows.append([
            "Due Date",
            "Customer Name",
            "Part Number",
            "Job Number",
            "EXT",
            "Alloy",
            "Mold Type",
            "Quantity of Molds",
            "Castings Per Mold",
            "Cores Per Mold",
            "Total Weight per EXT",
            "# of Molds for EXT"
        ])

        for _, row in block["rows"].iterrows():

            excel_rows.append([
                row.get(Columns.COL_DUE_DATE, ""),
                row.get("Customer Name", ""),
                row.get("Part Number", ""),
                row.get(Columns.COL_JOB_NUMBER, ""),
                row.get("EXT", ""),
                row.get(Columns.COL_ALLOY, ""),
                row.get(Columns.COL_CAST_TYPE, ""),
                row.get("Quantity of Molds", ""),
                row.get("Castings Per Mold", ""),
                row.get("Quantity of Cores", ""),
                row.get("Total Weight per EXT", ""),
                row.get("Molds for EXT", "")
            ])

        excel_rows.append([
            "TOTALS",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            block["weight_total"],
            block["mold_total"]
        ])

        excel_rows.append([])

    return excel_rows

def Export_Mold_Schedule(
    Export_Blocks,
    output_file="Mold Schedule.xlsx"
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Mold Schedule"

    current_row = 1

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    bold = Font(bold=True)

    for day in sorted(Export_Blocks.keys()):

        block = Export_Blocks[day]

        # =====================================
        # Title Row
        # =====================================

        ws.cell(
            current_row,
            1,
            "Mold Schedule"
        )

        ws.cell(
            current_row,
            2,
            block["date"].strftime("%m/%d/%Y")
        )

        ws.cell(
            current_row,
            4,
            block["weekday"]
        )

        for col in range(1, 13):
            ws.cell(current_row, col).font = bold

        current_row += 2

        # =====================================
        # Headers
        # =====================================

        headers = [
            "Due Date",
            "Customer Name",
            "Part Number",
            "Job Number",
            "EXT",
            "Alloy",
            "Mold Type",
            "Quantity of Molds",
            "Castings Per Mold",
            "Cores Per Mold",
            "Total Weight per EXT",
            "# of Molds for EXT"
        ]

        for col_num, header in enumerate(
            headers,
            start=1
        ):
            cell = ws.cell(
                current_row,
                col_num,
                header
            )

            cell.font = bold
            cell.border = thin

        current_row += 1

        # =====================================
        # Data Rows
        # =====================================

        for _, row in block["rows"].iterrows():

            values = [
                row.get(Columns.COL_DUE_DATE, ""),
                row.get("Customer Name", ""),
                row.get("Part Number", ""),
                row.get(Columns.COL_JOB_NUMBER, ""),
                row.get("EXT", ""),
                row.get(Columns.COL_ALLOY, ""),
                row.get(config.COL_CAST_TYPE, ""),
                row.get("Quantity of Molds", ""),
                row.get("Castings Per Mold", ""),
                row.get("Quantity of Cores", ""),
                row.get("Total Weight per EXT", ""),
                row.get("Molds for EXT", "")
            ]

            for col_num, value in enumerate(
                values,
                start=1
            ):
                cell = ws.cell(
                    current_row,
                    col_num,
                    value
                )

                cell.border = thin

            current_row += 1

        # =====================================
        # Totals
        # =====================================

        ws.cell(
            current_row,
            1,
            "TOTALS"
        )

        ws.cell(
            current_row,
            11,
            block["weight_total"]
        )

        ws.cell(
            current_row,
            12,
            block["mold_total"]
        )

        ws.cell(current_row, 1).font = bold
        ws.cell(current_row, 11).font = bold
        ws.cell(current_row, 12).font = bold

        current_row += 3

    # =====================================
    # Widths
    # =====================================

    widths = {
        "A": 12,
        "B": 30,
        "C": 25,
        "D": 15,
        "E": 6,
        "F": 15,
        "G": 12,
        "H": 15,
        "I": 15,
        "J": 15,
        "K": 18,
        "L": 15
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_file)

    print(f"Saved: {output_file}")

# // =======================================================
# // Scheduling Module - Prep and Mold
# // =======================================================

def Schedule_Molds():
    InputFile = Read_File()

    Jobs_to_schedule = Mold_Scheduler(InputFile)

    Schedule_rows = Build_Schedule_Rows(Jobs_to_schedule)

    # print(filtered_job_counts)

    Schedule_Data_Frame = pd.DataFrame(Schedule_rows)

    Schedule_Data_Frame = (
    Schedule_Data_Frame
    .sort_values(
        by=[
            config.COL_ALLOY,
            config.COL_DUE_DATE,
            config.COL_JOB_NUMBER,
            "Extension_Seq"
        ],
        ascending=[
            False,
            False,
            False,
            True
        ]
    )
    .reset_index(drop=True)
    )

    Schedule_Data_Frame = Assign_days(Schedule_Data_Frame)

    print("\nDay Totals")

    print(
        Schedule_Data_Frame.groupby("Schedule Day")
        ["Molds for EXT"]
        .sum()
    )

    print(
        Schedule_Data_Frame[
            [
                config.COL_JOB_NUMBER,
                "EXT",
                config.COL_ALLOY,
                "Molds for EXT",
                "Schedule Day"
            ]
        ]
    )

    print_bucket(Schedule_Data_Frame)

    Daily_Schedules = Build_Daily_Schedules(
        Schedule_Data_Frame
    )

    Day_Dates = Build_Schedule_Dates(
        Daily_Schedules,
        datetime.today() + timedelta(days=1)
    )

    Export_Blocks = Build_Daily_Export_Blocks(
        Daily_Schedules,
        Day_Dates
    )

    Print_Export_Blocks(
        Export_Blocks
    )

    return Export_Blocks





# // =======================================================
# // Scheduling Module - Melt
# // =======================================================
# def MeltScheduler():
#     for i in MoldsCompleted:
#         if #cell == 0 molds left:
#             # put in list for pouring
#         else continue
            
# // If job has been scheduled and mold shows complete, but no heat number assigned. Show ready to pour

# // If ready to pour, read into list

# // for item in list add to ready-to-pour schedule

# // =======================================================
# // Cleaning Schedule
# // =======================================================
# def CleanScheduler():
#     for i in CastingsCreated():
#         if #cell == heat number
#             # put in list for pouring
#         else continue

# If heat number assigned, but not shipped, show on cleaning schedule sorted by due date.

# =======================================================
# Main Program Entry
# =======================================================

Export_Blocks = Schedule_Molds()

output_file = (
    r"C:\Users\lburkardt\OneDrive - MonettMetalsUS1"
    r"\Quality\Schedule\Output\Mold Schedule.xlsx"
)

Export_Mold_Schedule(
    Export_Blocks,
    output_file
)