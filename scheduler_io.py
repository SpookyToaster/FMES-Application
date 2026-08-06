"""
File I/O helpers for the mold production scheduler.

Currently handles reading the Open Order Report Excel workbook that is
exported from the ERP system and placed in the shared OneDrive folder.
"""

from datetime import datetime
from pathlib import Path
import shutil
import tempfile
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
import pandas as pd

from config import Columns
from DB_IO import get_main_dashboard_rows


DEFAULT_OPEN_ORDER_REPORT_PATH = (
    r"C:\Users\lburkardt\OneDrive - MonettMetalsUS1"
    r"\Quality\Schedule\Open Order Report.xlsx"
)

DEFAULT_BACKUP_DIR = (
    r"C:\Users\lburkardt\OneDrive - MonettMetalsUS1"
    r"\Quality\Schedule\Backups"
)

DEFAULT_HISTORICAL_OOR_DIR = (
    r"C:\Users\lburkardt\OneDrive - MonettMetalsUS1"
    r"\Quality\Schedule\Historical OORs"
)

DEFAULT_DB_SNAPSHOT_DIR = (
    r"C:\Users\lburkardt\OneDrive - MonettMetalsUS1"
    r"\Quality\Schedule\Historical DB Snapshots"
)

SQL_MAIN_EXPORT_COLUMNS = [
    "Due Date",
    "Customer Name",
    "Part Number",
    "Job Type",
    "Job Number",
    "Alloy",
    "Casting Type",
    "QTY Ordered",
    "Quantity of Molds",
    "Castings Per Mold",
    "Quantity of Cores",
    "Pour Weight",
    "Total Pour WT",
    "Total Value",
    "Heat No Assigned",
    "Castings Produced",
    "Molds Completed",
]

EXCLUDED_CUSTOMER_NAMES = {"MONETT"}
XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XR_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
XR2_NS = "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"
XR3_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
REQUIRED_SQL_TEXT_COLUMNS = [
    "Due Date",
    "Customer Name",
    "Part Number",
    "Job Number",
]


def _ensure_directory(path):
    """Create path (and parents) when missing."""
    Path(path).mkdir(parents=True, exist_ok=True)


def _exclude_rows_by_customer_name(rows, excluded_names=EXCLUDED_CUSTOMER_NAMES):
    """Return rows excluding records whose Customer Name matches excluded_names."""
    normalized_exclusions = {str(name).strip().upper() for name in excluded_names}

    filtered = []
    for row in rows:
        customer_name = str(row.get("Customer Name", "")).strip().upper()
        if customer_name in normalized_exclusions:
            continue
        filtered.append(row)

    return filtered


def _validate_sql_rows(rows):
    """
    Validate that SQL rows are complete enough for overwrite and scheduling.

    Raises RuntimeError with a concise defect summary when required fields are
    missing so bad joins/mappings do not silently create invalid schedules.
    """
    if not rows:
        raise RuntimeError(
            "SQL scheduler input validation failed. No rows were returned after SQL join/mapping. "
            "No matching Job Number keys were found between the source datasets."
        )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return rows

    issues = []
    for column_name in REQUIRED_SQL_TEXT_COLUMNS:
        if column_name not in frame.columns:
            issues.append(f"missing column: {column_name}")
            continue

        series = frame[column_name]
        invalid_mask = series.isna() | (series.astype(str).str.strip() == "")
        invalid_count = int(invalid_mask.sum())
        if invalid_count:
            sample_jobs = []
            if "Job Number" in frame.columns:
                sample_jobs = (
                    frame.loc[invalid_mask, "Job Number"]
                    .astype(str)
                    .head(5)
                    .tolist()
                )
            issues.append(
                f"{column_name}: {invalid_count} blank rows"
                + (f" (sample jobs: {', '.join(sample_jobs)})" if sample_jobs else "")
            )

    if issues:
        raise RuntimeError(
            "SQL scheduler input validation failed. The live query is returning incomplete rows: "
            + "; ".join(issues)
            + ". Provide the Power Query relationship/join logic so the SQL source can be mapped correctly."
        )

    return rows


def _to_plain_text(value):
    """Return plain-text cell value for workbook writes without touching formatting."""
    if pd.isna(value):
        return ""

    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%Y-%m-%d")

    return str(value)


def _next_incremented_path(directory, stem, suffix=".xlsx"):
    """Return first available path using an incremented numeric suffix."""
    directory = Path(directory)
    index = 1

    while True:
        candidate = directory / f"{stem}_{index:03d}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def _excel_column_letter(column_index):
    """Return Excel column letters for a 1-based column index."""
    letters = []
    while column_index > 0:
        column_index, remainder = divmod(column_index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _excel_column_index(column_letters):
    """Return 1-based numeric column index for Excel column letters."""
    value = 0
    for char in column_letters:
        value = (value * 26) + (ord(char.upper()) - 64)
    return value


def _sheet_cell_reference(row_index, column_index):
    """Return Excel-style cell reference like F2."""
    return f"{_excel_column_letter(column_index)}{row_index}"


def _cell_column_index(cell_reference):
    """Extract numeric column index from an Excel cell reference."""
    letters = "".join(char for char in cell_reference if char.isalpha())
    return _excel_column_index(letters)


def _resolve_sheet_archive_path(zip_file, sheet_name):
    """Resolve workbook sheet name to xl/worksheets/... archive path."""
    workbook_root = ET.fromstring(zip_file.read("xl/workbook.xml"))
    relationships_root = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))

    relationship_id = None
    for sheet in workbook_root.findall(f"{{{XML_NS}}}sheets/{{{XML_NS}}}sheet"):
        if sheet.get("name") == sheet_name:
            relationship_id = sheet.get(f"{{{REL_NS}}}id")
            break

    if not relationship_id:
        raise RuntimeError(f"Worksheet '{sheet_name}' was not found.")

    for relationship in relationships_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
        if relationship.get("Id") == relationship_id:
            target = relationship.get("Target", "")
            if target.startswith("/"):
                return target.lstrip("/")
            return f"xl/{target.lstrip('/')}"

    raise RuntimeError(f"Worksheet relationship for '{sheet_name}' was not found.")


def _find_or_create_row(sheet_data, row_index):
    """Return existing row element or insert a new one in sorted position."""
    row_tag = f"{{{XML_NS}}}row"

    for row in sheet_data.findall(row_tag):
        if int(row.get("r", "0")) == row_index:
            return row

    new_row = ET.Element(row_tag, {"r": str(row_index)})
    inserted = False
    for position, row in enumerate(sheet_data.findall(row_tag)):
        if int(row.get("r", "0")) > row_index:
            sheet_data.insert(position, new_row)
            inserted = True
            break

    if not inserted:
        sheet_data.append(new_row)

    return new_row


def _find_or_create_cell(row_element, row_index, column_index):
    """Return existing cell element or insert a new one in sorted position."""
    cell_ref = _sheet_cell_reference(row_index, column_index)
    cell_tag = f"{{{XML_NS}}}c"

    for cell in row_element.findall(cell_tag):
        if cell.get("r") == cell_ref:
            return cell

    new_cell = ET.Element(cell_tag, {"r": cell_ref})
    inserted = False
    for position, cell in enumerate(row_element.findall(cell_tag)):
        if _cell_column_index(cell.get("r", "A1")) > column_index:
            row_element.insert(position, new_cell)
            inserted = True
            break

    if not inserted:
        row_element.append(new_cell)

    return new_cell


def _set_cell_plain_text(cell, text):
    """Replace cell contents with inline plain text while preserving attributes/styles."""
    for child in list(cell):
        cell.remove(child)

    cell.set("t", "inlineStr")
    is_node = ET.SubElement(cell, f"{{{XML_NS}}}is")
    text_node = ET.SubElement(is_node, f"{{{XML_NS}}}t")
    text_node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    text_node.text = text


def _restore_ignorable_namespace_declarations(xml_bytes):
    """
    Restore namespace declarations referenced by Ignorable after ET serialization.

    ElementTree can rename/drop prefixes for unused namespaces, but worksheets may
    still include an Ignorable list that references those prefixes. Excel then
    repairs the sheet as malformed. This keeps the needed declarations present.
    """
    xml_text = xml_bytes.decode("utf-8")
    xml_decl_end = xml_text.find("?>")
    search_start = xml_decl_end + 2 if xml_decl_end >= 0 else 0
    root_tag_start = xml_text.find("<", search_start)
    if root_tag_start < 0:
        return xml_bytes

    root_tag_end = xml_text.find(">", root_tag_start)
    if root_tag_end <= root_tag_start:
        return xml_bytes

    root_open = xml_text[root_tag_start:root_tag_end]
    if "Ignorable=\"x14ac xr xr2 xr3\"" not in root_open:
        return xml_bytes

    missing_decls = []
    required = {
        "x14ac": X14AC_NS,
        "xr": XR_NS,
        "xr2": XR2_NS,
        "xr3": XR3_NS,
    }

    for prefix, uri in required.items():
        marker = f'xmlns:{prefix}="'
        if marker not in root_open:
            missing_decls.append(f' xmlns:{prefix}="{uri}"')

    if not missing_decls:
        return xml_bytes

    patched = xml_text[:root_tag_end] + "".join(missing_decls) + xml_text[root_tag_end:]
    return patched.encode("utf-8")


def _export_worksheet_values(source_workbook_path, sheet_name, output_path):
    """Export raw worksheet values into a standalone workbook."""
    source_wb = load_workbook(source_workbook_path, data_only=True)
    try:
        if sheet_name not in source_wb.sheetnames:
            raise RuntimeError(f"Worksheet '{sheet_name}' was not found.")

        source_ws = source_wb[sheet_name]

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        for row in source_ws.iter_rows(
            min_row=1,
            max_row=source_ws.max_row,
            min_col=1,
            max_col=source_ws.max_column,
            values_only=True,
        ):
            new_ws.append(list(row))

        new_wb.save(output_path)
    finally:
        source_wb.close()


def _write_sql_data_to_oor(source_workbook_path, sql_rows, sheet_name="OOR"):
    """Overwrite OOR F:V values by editing sheet XML directly to preserve workbook metadata."""
    start_row = 2
    start_col = 6   # F
    end_col = 22    # V

    with zipfile.ZipFile(source_workbook_path, "r") as source_zip:
        sheet_archive_path = _resolve_sheet_archive_path(source_zip, sheet_name)
        sheet_root = ET.fromstring(source_zip.read(sheet_archive_path))
        sheet_data = sheet_root.find(f"{{{XML_NS}}}sheetData")

        if sheet_data is None:
            raise RuntimeError(f"Worksheet '{sheet_name}' does not contain sheet data.")

        max_existing_row = 0
        for row in sheet_data.findall(f"{{{XML_NS}}}row"):
            max_existing_row = max(max_existing_row, int(row.get("r", "0")))

        if max_existing_row >= start_row:
            for row_idx in range(start_row, max_existing_row + 1):
                row_element = _find_or_create_row(sheet_data, row_idx)
                for col_idx in range(start_col, end_col + 1):
                    cell = _find_or_create_cell(row_element, row_idx, col_idx)
                    _set_cell_plain_text(cell, "")

        for offset, sql_row in enumerate(sql_rows):
            row_idx = start_row + offset
            row_element = _find_or_create_row(sheet_data, row_idx)
            for col_offset, col_name in enumerate(SQL_MAIN_EXPORT_COLUMNS):
                col_idx = start_col + col_offset
                cell = _find_or_create_cell(row_element, row_idx, col_idx)
                _set_cell_plain_text(cell, _to_plain_text(sql_row.get(col_name, "")))

        updated_sheet_xml = ET.tostring(
            sheet_root,
            encoding="utf-8",
            xml_declaration=True,
        )
        updated_sheet_xml = _restore_ignorable_namespace_declarations(updated_sheet_xml)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_path = temp_file.name

        try:
            with zipfile.ZipFile(temp_path, "w") as target_zip:
                for zip_info in source_zip.infolist():
                    if zip_info.filename == sheet_archive_path:
                        target_zip.writestr(zip_info, updated_sheet_xml)
                    else:
                        target_zip.writestr(zip_info, source_zip.read(zip_info.filename))
            shutil.move(temp_path, source_workbook_path)
        finally:
            if Path(temp_path).exists():
                Path(temp_path).unlink(missing_ok=True)


def _save_sql_snapshot(sql_rows, output_path):
    """Save SQL rows to a standalone snapshot workbook for day-to-day diffing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SQL Snapshot"

    ws.append(SQL_MAIN_EXPORT_COLUMNS)
    for row in sql_rows:
        ws.append([_to_plain_text(row.get(col, "")) for col in SQL_MAIN_EXPORT_COLUMNS])

    wb.save(output_path)


def Sync_Open_Order_Report_With_SQL(
    source_workbook_path=DEFAULT_OPEN_ORDER_REPORT_PATH,
    backup_dir=DEFAULT_BACKUP_DIR,
    historical_oor_dir=DEFAULT_HISTORICAL_OOR_DIR,
    db_snapshot_dir=DEFAULT_DB_SNAPSHOT_DIR,
    run_id=None,
    start_due_date=None,
    end_due_date=None,
):
    """
    Sync Open Order Report workbook artifacts with current SQL main-dashboard data.

    Steps:
      1) Backup source workbook to incremented file in backup_dir.
      2) Export current OOR worksheet values to historical OOR workbook.
      3) Overwrite OOR values in F2:V* using SQL rows as plain text.
      4) Save SQL rows to a dated snapshot workbook.

    Returns:
        dict with file paths and row_count written.
    """
    source_path = Path(source_workbook_path)
    if not source_path.exists():
        raise RuntimeError(f"Open Order Report workbook was not found at {source_workbook_path}")

    _ensure_directory(backup_dir)
    _ensure_directory(historical_oor_dir)
    _ensure_directory(db_snapshot_dir)

    timestamp = datetime.now()

    backup_path = _next_incremented_path(backup_dir, source_path.stem)
    shutil.copy2(source_path, backup_path)

    historical_oor_path = _next_incremented_path(
        historical_oor_dir,
        f"OOR-{timestamp.strftime('%Y-%m-%d')}",
    )
    _export_worksheet_values(source_path, "OOR", historical_oor_path)

    sql_rows = get_main_dashboard_rows(
        run_id=run_id,
        start_due_date=start_due_date,
        end_due_date=end_due_date,
    )
    sql_rows = _exclude_rows_by_customer_name(sql_rows)
    sql_rows = _validate_sql_rows(sql_rows)

    _write_sql_data_to_oor(source_path, sql_rows, sheet_name="OOR")

    db_snapshot_path = Path(db_snapshot_dir) / (
        f"DB-Snapshot-{timestamp.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    )
    _save_sql_snapshot(sql_rows, db_snapshot_path)

    return {
        "backup_path": str(backup_path),
        "historical_oor_path": str(historical_oor_path),
        "db_snapshot_path": str(db_snapshot_path),
        "row_count": len(sql_rows),
    }


def _coerce_numeric(series_like):
    """Return numeric Series with NaN converted to 0."""
    return pd.to_numeric(series_like, errors="coerce").fillna(0)


def _normalize_sql_rows(raw_rows):
    """
    Normalize SQL dashboard rows into the schema expected by the scheduler.

    The SQL source provides total molds and completed molds.  The scheduler
    expects "Molds Needed" as remaining molds, so we derive it as:
        max(Quantity of Molds - Molds Completed, 0)
    """
    frame = pd.DataFrame(raw_rows)

    if frame.empty:
        return frame

    frame.columns = frame.columns.str.strip()

    if "Molds Completed" not in frame.columns:
        frame["Molds Completed"] = 0

    if "Quantity of Molds" not in frame.columns:
        frame["Quantity of Molds"] = 0

    quantity_of_molds = _coerce_numeric(frame["Quantity of Molds"])
    molds_completed = _coerce_numeric(frame["Molds Completed"])

    frame["Quantity of Molds"] = quantity_of_molds
    frame["Molds Completed"] = molds_completed
    frame[Columns.COL_MOLDS_NEEDED] = (quantity_of_molds - molds_completed).clip(lower=0)

    if Columns.COL_HOLD not in frame.columns:
        frame[Columns.COL_HOLD] = "NO"
    else:
        frame[Columns.COL_HOLD] = frame[Columns.COL_HOLD].fillna("NO")

    if Columns.COL_SCHEDULED not in frame.columns:
        frame[Columns.COL_SCHEDULED] = "NO"
    else:
        frame[Columns.COL_SCHEDULED] = frame[Columns.COL_SCHEDULED].fillna("NO")

    for required_col in [
        Columns.COL_DUE_DATE,
        Columns.COL_JOB_NUMBER,
        Columns.COL_POUR_WEIGHT,
        Columns.COL_JOB_TYPE,
        Columns.COL_ALLOY,
        Columns.COL_CAST_TYPE,
        "Customer Name",
        "Part Number",
        "Castings Per Mold",
        "Quantity of Cores",
    ]:
        if required_col not in frame.columns:
            frame[required_col] = ""

    return frame


def Read_File(
    filepath=DEFAULT_OPEN_ORDER_REPORT_PATH,
    source="excel",
    run_id=None,
    start_due_date=None,
    end_due_date=None,
):
    """
    Read scheduler input from either Excel or SQL and return it as a DataFrame.

    Strips leading/trailing whitespace from column headers so downstream
    column lookups are not affected by inconsistent ERP exports.

    Args:
        filepath: Excel path used when source="excel".
        source:   "excel" or "sql".
        run_id:   Optional SchedulerRun.RunId used when source="sql".
        start_due_date: Optional inclusive lower bound (source="sql").
        end_due_date:   Optional inclusive upper bound (source="sql").

    Returns:
        pandas DataFrame with scheduler-compatible columns.

    Raises:
        RuntimeError: If source read fails or source is invalid.
    """
    try:
        if source == "excel":
            imported_file = pd.read_excel(filepath, sheet_name="OOR")
            imported_file.columns = imported_file.columns.str.strip()
            return imported_file

        if source == "sql":
            raw_rows = get_main_dashboard_rows(
                run_id=run_id,
                start_due_date=start_due_date,
                end_due_date=end_due_date,
            )
            raw_rows = _exclude_rows_by_customer_name(raw_rows)
            raw_rows = _validate_sql_rows(raw_rows)
            return _normalize_sql_rows(raw_rows)

        raise RuntimeError(f"Unsupported input source '{source}'. Use 'excel' or 'sql'.")
    except Exception as exc:
        if source == "excel":
            raise RuntimeError(
                f"Failed to read schedule input from {filepath} (sheet 'OOR')"
            ) from exc

        if source == "sql":
            raise RuntimeError(
                f"Failed to read schedule input from SQL main dashboard query: {exc}"
            ) from exc

        raise
