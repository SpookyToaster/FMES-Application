import sys
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scheduler_io import (
    Can_Alloy_Share_Heat_With,
    Read_File,
    SQL_MAIN_EXPORT_COLUMNS,
    Sync_Open_Order_Report_With_SQL,
    _restore_ignorable_namespace_declarations,
)


class SchedulerIOTests(unittest.TestCase):
    def test_restore_ignorable_namespace_declarations_adds_missing_prefixes(self):
        xml_bytes = (
            b'<?xml version="1.0" encoding="utf-8"?>'
            b'<s:worksheet xmlns:s="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
            b'mc:Ignorable="x14ac xr xr2 xr3"><s:sheetData/></s:worksheet>'
        )

        patched = _restore_ignorable_namespace_declarations(xml_bytes).decode("utf-8")

        self.assertIn('xmlns:x14ac="http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"', patched)
        self.assertIn('xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision"', patched)
        self.assertIn('xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"', patched)
        self.assertIn('xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"', patched)

    def test_read_file_strips_headers(self):
        frame = pd.DataFrame(columns=[" Job Number ", " Due Date "])

        with patch("scheduler_io.pd.read_excel", return_value=frame):
            result = Read_File("sample.xlsx")

        self.assertEqual(list(result.columns)[:2], ["Job Number", "Due Date"])
        self.assertIn("Compatibility Group", result.columns)
        self.assertIn("Compatible With ASTM Group", result.columns)
        self.assertIn("Specific Compatible Alloys", result.columns)

    def test_read_file_wraps_failures(self):
        with patch("scheduler_io.pd.read_excel", side_effect=FileNotFoundError("missing")):
            with self.assertRaises(RuntimeError) as context:
                Read_File(filepath="missing.xlsx", source="excel")

        self.assertIn("Failed to read schedule input", str(context.exception))

    def test_read_file_sql_normalizes_required_scheduler_columns(self):
        sql_rows = [
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Customer",
                "Part Number": "P1",
                "Job Type": "JOB",
                "Job Number": "9001",
                "Alloy": "A",
                "Casting Type": "L",
                "Quantity of Molds": 10,
                "Castings Per Mold": 2,
                "Quantity of Cores": 1,
                "Pour Weight": 100,
                "Molds Completed": 3,
            }
        ]

        with patch("scheduler_io.get_main_dashboard_rows", return_value=sql_rows):
            frame = Read_File(source="sql")

        self.assertEqual(frame.iloc[0]["Molds Needed"], 7)
        self.assertEqual(frame.iloc[0]["Hold"], "NO")
        self.assertEqual(frame.iloc[0]["Scheduled"], "NO")

    def test_read_file_sql_wraps_failures(self):
        with patch("scheduler_io.get_main_dashboard_rows", side_effect=RuntimeError("db down")):
            with self.assertRaises(RuntimeError) as context:
                Read_File(source="sql")

        self.assertIn("Failed to read schedule input from SQL", str(context.exception))

    def test_read_file_sql_excludes_monett_customer_rows(self):
        sql_rows = [
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Monett",
                "Part Number": "P-MON",
                "Job Type": "JOB",
                "Job Number": "M-1",
                "Alloy": "A",
                "Casting Type": "L",
                "Quantity of Molds": 10,
                "Molds Completed": 0,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Pour Weight": 100,
            },
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Customer B",
                "Part Number": "P-OK",
                "Job Type": "JOB",
                "Job Number": "B-1",
                "Alloy": "A",
                "Casting Type": "L",
                "Quantity of Molds": 8,
                "Molds Completed": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Pour Weight": 90,
            },
        ]

        with patch("scheduler_io.get_main_dashboard_rows", return_value=sql_rows):
            frame = Read_File(source="sql")

        self.assertEqual(len(frame), 1)
        self.assertEqual(frame.iloc[0]["Customer Name"], "Customer B")

    def test_read_file_sql_reports_incomplete_joined_rows(self):
        sql_rows = [
            {
                "Due Date": "",
                "Customer Name": "Customer B",
                "Part Number": "",
                "Job Type": "",
                "Job Number": "B-1",
                "Alloy": "",
                "Casting Type": "",
                "Quantity of Molds": 8,
                "Molds Completed": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Pour Weight": 90,
            },
        ]

        with patch("scheduler_io.get_main_dashboard_rows", return_value=sql_rows):
            with self.assertRaises(RuntimeError) as context:
                Read_File(source="sql")

        message = str(context.exception)
        self.assertIn("SQL scheduler input validation failed", message)
        self.assertIn("Due Date", message)
        self.assertIn("Part Number", message)

    def test_read_file_sql_reports_zero_matched_rows(self):
        with patch("scheduler_io.get_main_dashboard_rows", return_value=[]):
            with self.assertRaises(RuntimeError) as context:
                Read_File(source="sql")

        self.assertIn("No rows were returned", str(context.exception))

    def test_read_file_sql_allows_blank_alloy_rows(self):
        sql_rows = [
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Customer A",
                "Part Number": "P1",
                "Job Type": "JOB",
                "Job Number": "VUPN",
                "Alloy": "",
                "Casting Type": "L",
                "Quantity of Molds": 5,
                "Molds Completed": 1,
                "Castings Per Mold": 2,
                "Quantity of Cores": 1,
                "Pour Weight": 100,
            },
        ]

        with patch("scheduler_io.get_main_dashboard_rows", return_value=sql_rows):
            frame = Read_File(source="sql")

        self.assertEqual(len(frame), 1)
        self.assertEqual(frame.iloc[0]["Job Number"], "VUPN")
        self.assertEqual(frame.iloc[0]["Alloy"], "")

    def test_read_file_sql_applies_alloy_compatibility_columns(self):
        sql_rows = [
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Customer A",
                "Part Number": "P1",
                "Job Type": "JOB",
                "Job Number": "9001",
                "Alloy": "WCB",
                "Casting Type": "L",
                "Quantity of Molds": 5,
                "Molds Completed": 1,
                "Castings Per Mold": 2,
                "Quantity of Cores": 1,
                "Pour Weight": 100,
            },
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Customer B",
                "Part Number": "P2",
                "Job Type": "JOB",
                "Job Number": "9002",
                "Alloy": "4140",
                "Casting Type": "L",
                "Quantity of Molds": 6,
                "Molds Completed": 2,
                "Castings Per Mold": 2,
                "Quantity of Cores": 1,
                "Pour Weight": 120,
            },
        ]

        compatibility_frame = pd.DataFrame(
            [
                {
                    "alloy_code": "WCB",
                    "compatibility_group": "A216",
                    "Is_Compat_with_All": "YES",
                },
                {
                    "alloy_code": "4140",
                    "compatibility_group": "4140",
                    "Is_Compat_with_All": "NO",
                },
            ]
        )

        with patch("scheduler_io.get_main_dashboard_rows", return_value=sql_rows), patch(
            "scheduler_io.pd.read_csv", return_value=compatibility_frame
        ):
            frame = Read_File(source="sql")

        self.assertEqual(frame.iloc[0]["Compatibility Group"], "A216")
        self.assertEqual(frame.iloc[0]["Compatible With ASTM Group"], "YES")
        self.assertEqual(frame.iloc[1]["Compatibility Group"], "4140")
        self.assertEqual(frame.iloc[1]["Compatible With ASTM Group"], "NO")
        self.assertIn("Specific Compatible Alloys", frame.columns)

    def test_read_file_sql_preserves_specific_directional_compatibility(self):
        sql_rows = [
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Customer A",
                "Part Number": "P1",
                "Job Type": "JOB",
                "Job Number": "9001",
                "Alloy": "CF8M",
                "Casting Type": "L",
                "Quantity of Molds": 5,
                "Molds Completed": 1,
                "Castings Per Mold": 2,
                "Quantity of Cores": 1,
                "Pour Weight": 100,
            },
            {
                "Due Date": "2026-08-04",
                "Customer Name": "Customer B",
                "Part Number": "P2",
                "Job Type": "JOB",
                "Job Number": "9002",
                "Alloy": "CF3M",
                "Casting Type": "L",
                "Quantity of Molds": 6,
                "Molds Completed": 2,
                "Castings Per Mold": 2,
                "Quantity of Cores": 1,
                "Pour Weight": 120,
            },
        ]

        compatibility_frame = pd.DataFrame(
            [
                {
                    "alloy_code": "CF8M",
                    "compatibility_group": "A351",
                    "Is_Compat_with_All": "NO",
                    "compatible_specific_alloys": "CF3M",
                },
                {
                    "alloy_code": "CF3M",
                    "compatibility_group": "A351",
                    "Is_Compat_with_All": "NO",
                    "compatible_specific_alloys": "",
                },
            ]
        )

        with patch("scheduler_io.get_main_dashboard_rows", return_value=sql_rows), patch(
            "scheduler_io.pd.read_csv", return_value=compatibility_frame
        ):
            frame = Read_File(source="sql")

        self.assertEqual(frame.iloc[0]["Compatibility Group"], "A351")
        self.assertEqual(frame.iloc[0]["Specific Compatible Alloys"], "CF3M")
        self.assertEqual(frame.iloc[1]["Compatibility Group"], "A351")
        self.assertEqual(frame.iloc[1]["Specific Compatible Alloys"], "")

    def test_can_alloy_share_heat_with_allows_groupwide_astm_compatibility(self):
        compatibility_map = {
            "80-40": {
                "Compatibility Group": "A148",
                "Compatible With ASTM Group": "YES",
                "Specific Compatible Alloys": "",
            },
            "150-135": {
                "Compatibility Group": "A148",
                "Compatible With ASTM Group": "YES",
                "Specific Compatible Alloys": "",
            },
        }

        self.assertTrue(
            Can_Alloy_Share_Heat_With("80-40", "150-135", compatibility_map=compatibility_map)
        )

    def test_can_alloy_share_heat_with_honors_directional_specific_compatibility(self):
        compatibility_map = {
            "CF3M": {
                "Compatibility Group": "A351",
                "Compatible With ASTM Group": "NO",
                "Specific Compatible Alloys": "CF8M",
            },
            "CF8M": {
                "Compatibility Group": "A351",
                "Compatible With ASTM Group": "NO",
                "Specific Compatible Alloys": "",
            },
        }

        self.assertTrue(
            Can_Alloy_Share_Heat_With("CF3M", "CF8M", compatibility_map=compatibility_map)
        )
        self.assertFalse(
            Can_Alloy_Share_Heat_With("CF8M", "CF3M", compatibility_map=compatibility_map)
        )

    def test_sync_open_order_report_with_sql_writes_backup_history_and_snapshot(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "Open Order Report.xlsx"
            backup_dir = temp_path / "Backups"
            hist_dir = temp_path / "Historical OORs"
            snap_dir = temp_path / "Historical DB Snapshots"

            workbook = Workbook()
            ws = workbook.active
            ws.title = "OOR"

            for idx in range(1, 23):
                ws.cell(row=1, column=idx).value = f"H{idx}"

            ws.cell(row=2, column=6).value = "old-f2"
            ws.cell(row=3, column=6).value = "old-f3"
            workbook.save(source_path)

            sql_rows = [
                {
                    "Due Date": "2026-08-04",
                    "Customer Name": "Monett",
                    "Part Number": "P-MON",
                    "Job Type": "JOB",
                    "Job Number": "M-1",
                    "Alloy": "A",
                    "Casting Type": "L",
                    "QTY Ordered": 3,
                    "Quantity of Molds": 3,
                    "Castings Per Mold": 1,
                    "Quantity of Cores": 0,
                    "Pour Weight": 60,
                    "Total Pour WT": 180,
                    "Total Value": 100,
                    "Heat No Assigned": "HM",
                    "Castings Produced": 0,
                    "Molds Completed": 0,
                },
                {
                    "Due Date": "2026-08-04",
                    "Customer Name": "Customer A",
                    "Part Number": "P1",
                    "Job Type": "JOB",
                    "Job Number": "9001",
                    "Alloy": "A",
                    "Casting Type": "L",
                    "QTY Ordered": 10,
                    "Quantity of Molds": 5,
                    "Castings Per Mold": 2,
                    "Quantity of Cores": 1,
                    "Pour Weight": 100,
                    "Total Pour WT": 500,
                    "Total Value": 1000,
                    "Heat No Assigned": "H1",
                    "Castings Produced": 3,
                    "Molds Completed": 1,
                }
            ]

            with patch("scheduler_io.get_main_dashboard_rows", return_value=sql_rows):
                result = Sync_Open_Order_Report_With_SQL(
                    source_workbook_path=str(source_path),
                    backup_dir=str(backup_dir),
                    historical_oor_dir=str(hist_dir),
                    db_snapshot_dir=str(snap_dir),
                )

            self.assertEqual(result["row_count"], 1)
            self.assertTrue(Path(result["backup_path"]).exists())
            self.assertTrue(Path(result["historical_oor_path"]).exists())
            self.assertTrue(Path(result["db_snapshot_path"]).exists())

            synced_wb = load_workbook(source_path)
            synced_ws = synced_wb["OOR"]
            self.assertEqual(synced_ws.cell(row=2, column=6).value, "2026-08-04")
            self.assertEqual(synced_ws.cell(row=2, column=7).value, "Customer A")
            self.assertEqual(synced_ws.cell(row=2, column=13).value, 10)
            self.assertEqual(synced_ws.cell(row=2, column=14).value, 5)
            self.assertEqual(synced_ws.cell(row=2, column=15).value, 2)
            self.assertEqual(synced_ws.cell(row=2, column=16).value, 1)
            self.assertEqual(synced_ws.cell(row=2, column=17).value, 100)
            self.assertEqual(synced_ws.cell(row=2, column=18).value, 500)
            self.assertEqual(synced_ws.cell(row=2, column=19).value, 1000)
            self.assertEqual(synced_ws.cell(row=2, column=21).value, 3)
            self.assertEqual(synced_ws.cell(row=2, column=22).value, "1")
            self.assertEqual(synced_ws.cell(row=3, column=6).value, "")
            synced_wb.close()

            snapshot_wb = load_workbook(result["db_snapshot_path"])
            snapshot_ws = snapshot_wb["SQL Snapshot"]
            header_values = [snapshot_ws.cell(row=1, column=i + 1).value for i in range(len(SQL_MAIN_EXPORT_COLUMNS))]
            self.assertEqual(header_values, SQL_MAIN_EXPORT_COLUMNS)
            snapshot_wb.close()


if __name__ == "__main__":
    unittest.main()