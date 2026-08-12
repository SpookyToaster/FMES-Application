import sys
from datetime import datetime
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from fmes.config import Columns
from fmes.scheduler_export import build_daily_export_blocks, build_excel_rows, build_heat_daily_totals_rows, build_heat_detail_rows, build_heat_planner_rows, build_heat_summary_rows, build_job_shipping_report_rows, export_combined_schedule_workbook, export_heat_summary, export_mold_schedule


class SchedulerExportTests(unittest.TestCase):
    def test_build_daily_export_blocks_sorts_rows_by_alloy(self):
        frame = pd.DataFrame([
            {
                Columns.COL_DUE_DATE: "2026-08-05",
                "Customer Name": "Customer B",
                "Part Number": "P2",
                Columns.COL_JOB_NUMBER: "5002",
                "EXT": "",
                Columns.COL_ALLOY: "B",
                Columns.COL_CAST_TYPE: "L",
                "Quantity of Molds": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Total Weight per EXT": 30,
                "Molds for EXT": 1,
            },
            {
                Columns.COL_DUE_DATE: "2026-08-04",
                "Customer Name": "Customer A",
                "Part Number": "P1",
                Columns.COL_JOB_NUMBER: "5001",
                "EXT": "",
                Columns.COL_ALLOY: "A",
                Columns.COL_CAST_TYPE: "L",
                "Quantity of Molds": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Total Weight per EXT": 25,
                "Molds for EXT": 1,
            },
        ])

        blocks = build_daily_export_blocks(
            {1: frame},
            {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}},
        )

        ordered_alloys = blocks[1]["rows"][Columns.COL_ALLOY].tolist()
        self.assertEqual(ordered_alloys, ["A", "B"])

    def test_build_export_blocks_and_excel_rows(self):
        frame = pd.DataFrame([
            {
                Columns.COL_DUE_DATE: "2026-08-04",
                "Customer Name": "Customer",
                "Part Number": "P1",
                Columns.COL_JOB_NUMBER: "5001",
                "EXT": "",
                Columns.COL_ALLOY: "A",
                Columns.COL_CAST_TYPE: "L",
                "Quantity of Molds": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Total Weight per EXT": 25,
                "Molds for EXT": 2,
            }
        ])

        daily_schedules = {1: frame}
        day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}

        blocks = build_daily_export_blocks(daily_schedules, day_dates)
        excel_rows = build_excel_rows(blocks)

        self.assertEqual(blocks[1]["weight_total"], 25)
        self.assertEqual(blocks[1]["mold_total"], 2)
        self.assertGreater(len(excel_rows), 0)
        self.assertEqual(excel_rows[1][-1], "Number of Molds Today")

    def test_export_mold_schedule_writes_file(self):
        frame = pd.DataFrame([
            {
                Columns.COL_DUE_DATE: "2026-08-04 00:00:00",
                "Customer Name": "Customer",
                "Part Number": "P1",
                Columns.COL_JOB_NUMBER: "5001",
                "EXT": "",
                Columns.COL_ALLOY: "A",
                Columns.COL_CAST_TYPE: "L",
                "Quantity of Molds": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Total Weight per EXT": 25,
                "Molds for EXT": 2,
                "Heat #": 1,
            }
        ])

        export_blocks = {
            1: {
                "date": pd.Timestamp("2026-08-04"),
                "weekday": "Tuesday",
                "rows": frame,
                "weight_total": 25,
                "mold_total": 2,
            },
            2: {
                "date": pd.Timestamp("2026-08-05"),
                "weekday": "Wednesday",
                "rows": pd.DataFrame([
                    {
                        Columns.COL_DUE_DATE: "2026-08-05 00:00:00",
                        "Customer Name": "Customer Two",
                        "Part Number": "P2",
                        Columns.COL_JOB_NUMBER: "5002",
                        "EXT": "B",
                        Columns.COL_ALLOY: "B",
                        Columns.COL_CAST_TYPE: "N",
                        "Quantity of Molds": 1,
                        "Castings Per Mold": 1,
                        "Quantity of Cores": 0,
                        "Total Weight per EXT": 30,
                        "Molds for EXT": 1,
                        "Heat #": 2,
                    }
                ]),
                "weight_total": 30,
                "mold_total": 1,
            }
        }

        job_shipping_rows = [
            {
                "Job Number": "5001",
                "Schedule Status": "Scheduled",
                "Planned Molds": 2,
                "Scheduled Molds": 2,
                "Mold Day": 1,
                "Mold Date": pd.Timestamp("2026-08-04").date(),
                "Pour Day": 2,
                "Pour Date": pd.Timestamp("2026-08-05").date(),
                "Expected Ship Date": pd.Timestamp("2026-08-19").date(),
                "Due Date": pd.Timestamp("2026-08-20").date(),
                "Ship Buffer Days": 1,
                "On-Time": "YES",
            }
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "mold_schedule.xlsx"
            export_mold_schedule(export_blocks, str(output_file), job_shipping_rows=job_shipping_rows)
            self.assertTrue(output_file.exists())

            wb = load_workbook(output_file)
            ws = wb["Mold Schedule"]
            self.assertEqual(ws.cell(7, 1).value, "Customer Name")
            self.assertEqual(ws.cell(7, 8).value, "Quantity of Cores")
            self.assertEqual(ws.cell(7, 10).value, "Number of Molds Today")
            self.assertEqual(ws.cell(8, 1).value, "Customer")
            self.assertEqual(ws.cell(8, 10).value, 2)
            self.assertEqual(ws.cell(1, 1).value, "Mold Schedule")
            self.assertEqual(ws.cell(2, 1).value, "Date Created")
            self.assertEqual(ws.cell(2, 4).value, "Average Molds / Day")
            self.assertEqual(ws.cell(5, 1).value, "Mold Schedule")

            ws_jobs = wb["Job Shipping Outlook"]
            self.assertEqual(ws_jobs.cell(1, 2).value, "Schedule Status")
            self.assertEqual(ws_jobs.cell(2, 1).value, "5001")
            self.assertEqual(ws_jobs.cell(2, 12).value, "YES")

            ws_job_status = wb["Overall Job Status"]
            self.assertEqual(ws_job_status.cell(1, 1).value, "Job Number")
            self.assertEqual(ws_job_status.cell(1, 3).value, "Expected Ship Date")
            self.assertEqual(ws_job_status.cell(1, 4).value, "Due Date")
            self.assertEqual(ws_job_status.cell(1, 6).value, "On-Time")
            self.assertEqual(ws_job_status.cell(2, 1).value, "5001")
            self.assertEqual(ws_job_status.cell(2, 6).value, "YES")

    def test_export_mold_schedule_includes_melt_sheet_when_data_is_provided(self):
        frame = pd.DataFrame([
            {
                Columns.COL_DUE_DATE: "2026-08-04 00:00:00",
                "Customer Name": "Customer",
                "Part Number": "P1",
                Columns.COL_JOB_NUMBER: "5001",
                "EXT": "",
                Columns.COL_ALLOY: "A",
                Columns.COL_CAST_TYPE: "L",
                "Quantity of Molds": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Total Weight per EXT": 25,
                "Molds for EXT": 2,
                "Heat #": 42,
            }
        ])

        export_blocks = {
            1: {
                "date": pd.Timestamp("2026-08-04"),
                "weekday": "Tuesday",
                "rows": frame,
                "weight_total": 25,
                "mold_total": 2,
            }
        }
        melt_schedule = {
            1: {"rows": frame}
        }
        day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}

        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "mold_schedule.xlsx"
            export_mold_schedule(export_blocks, str(output_file), melt_schedule=melt_schedule, day_dates=day_dates)
            wb = load_workbook(output_file)
            self.assertIn("Melt Schedule", wb.sheetnames)
            self.assertEqual(wb["Melt Schedule"].cell(1, 1).value, "Melt Schedule")
            self.assertEqual(wb["Melt Schedule"].cell(5, 1).value, "Melt Schedule")
            self.assertEqual(wb["Melt Schedule"].cell(7, 1).value, "Pour Date")

    def test_build_job_shipping_report_rows_marks_not_yet_scheduled(self):
        schedule_data_frame = pd.DataFrame([
            {
                Columns.COL_JOB_NUMBER: "5001",
                Columns.COL_DUE_DATE: "2026-08-20",
                "Molds for EXT": 2,
            },
            {
                Columns.COL_JOB_NUMBER: "5002",
                Columns.COL_DUE_DATE: "2026-08-18",
                "Molds for EXT": 1,
            },
        ])

        mold_schedule_frame = pd.DataFrame([
            {
                Columns.COL_JOB_NUMBER: "5001",
                "Molds for EXT": 2,
                "Schedule Day": 1,
                "Pour Schedule Day": 2,
            }
        ])

        mold_day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}
        pour_day_dates = {2: {"date": pd.Timestamp("2026-08-05"), "weekday": "Wednesday"}}

        rows = build_job_shipping_report_rows(
            schedule_data_frame,
            mold_schedule_frame,
            mold_day_dates,
            pour_day_dates,
        )

        by_job = {row["Job Number"]: row for row in rows}
        self.assertEqual(by_job["5001"]["Schedule Status"], "Scheduled")
        self.assertEqual(by_job["5001"]["Expected Ship Date"].isoformat(), "2026-08-19")
        self.assertEqual(by_job["5001"]["Ship Buffer Days"], 1)
        self.assertEqual(by_job["5001"]["On-Time"], "YES")

        self.assertEqual(by_job["5002"]["Schedule Status"], "Not Yet Scheduled")
        self.assertEqual(by_job["5002"]["On-Time"], "NOT SCHEDULED")

    def test_export_mold_schedule_creates_missing_parent_directory(self):
        export_blocks = {
            1: {
                "date": pd.Timestamp("2026-08-04"),
                "weekday": "Tuesday",
                "rows": pd.DataFrame([
                    {
                        Columns.COL_DUE_DATE: "2026-08-04 00:00:00",
                        "Customer Name": "Customer",
                        "Part Number": "P1",
                        Columns.COL_JOB_NUMBER: "5001",
                        "EXT": "",
                        Columns.COL_ALLOY: "A",
                        Columns.COL_CAST_TYPE: "L",
                        "Quantity of Molds": 1,
                        "Castings Per Mold": 1,
                        "Quantity of Cores": 0,
                        "Total Weight per EXT": 25,
                        "Molds for EXT": 2,
                        "Heat #": 1,
                    }
                ]),
                "weight_total": 25,
                "mold_total": 2,
            }
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "nested" / "mold_schedule.xlsx"
            export_mold_schedule(export_blocks, str(output_file))
            self.assertTrue(output_file.exists())

    def test_export_mold_schedule_wraps_save_failures(self):
        export_blocks = {
            1: {
                "date": pd.Timestamp("2026-08-04"),
                "weekday": "Tuesday",
                "rows": pd.DataFrame(),
                "weight_total": 0,
                "mold_total": 0,
            }
        }

        with patch("fmes.scheduler_export.Workbook.save", side_effect=PermissionError("locked")):
            with self.assertRaises(RuntimeError) as context:
                export_mold_schedule(export_blocks, "locked.xlsx")

        self.assertIn("Failed while exporting mold schedule", str(context.exception))

    def test_build_heat_summary_rows_uses_melt_plan_summary(self):
        heat_summary = pd.DataFrame([
            {
                "Heat Slot": 1,
                "Heat #": 1,
                "Heat Status": "Planned",
                "Planning Priority": "Highest Priority",
                "Review Window": "Next 2 Weeks",
                "Anchor Alloy": "LEW15",
                "Compatibility Group": "A216",
                "Earliest Due Date": pd.Timestamp("2026-08-04").date(),
                "Latest Due Date": pd.Timestamp("2026-08-04").date(),
                "Total Weight (lbs)": 1200,
                "Total Molds": 4,
                "Rows in Heat": 2,
                "Jobs": "5001",
                "Extensions": "5001-A, 5001-B",
            },
            {
                "Heat Slot": 6,
                "Heat #": "",
                "Heat Status": "Reserved",
                "Planning Priority": "",
                "Review Window": "",
                "Anchor Alloy": "",
                "Compatibility Group": "",
                "Earliest Due Date": "",
                "Latest Due Date": "",
                "Total Weight (lbs)": 0,
                "Total Molds": 0,
                "Rows in Heat": 0,
                "Jobs": "",
                "Extensions": "",
            },
        ])

        melt_schedule = {
            1: {
                "heat_summary": heat_summary,
                "rows": pd.DataFrame([
                    {
                        Columns.COL_DUE_DATE: "2026-08-04",
                        Columns.COL_JOB_NUMBER: "5001",
                        "EXT": "A",
                        "Molds for EXT": 2,
                        "Heat #": 1,
                    },
                    {
                        Columns.COL_DUE_DATE: "2026-08-05",
                        Columns.COL_JOB_NUMBER: "5001",
                        "EXT": "B",
                        "Molds for EXT": 2,
                        "Heat #": 1,
                    },
                ]),
            }
        }
        day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}

        mold_schedule_frame = pd.DataFrame([
            {
                Columns.COL_JOB_NUMBER: "5001",
                "EXT": "A",
                "Schedule Day": 1,
                "Pour Schedule Day": 1,
                "Heat #": 1,
                "Molds for EXT": 2,
            },
            {
                Columns.COL_JOB_NUMBER: "5001",
                "EXT": "B",
                "Schedule Day": 1,
                "Pour Schedule Day": 1,
                "Heat #": 1,
                "Molds for EXT": 2,
            },
        ])

        summary = build_heat_summary_rows((melt_schedule, day_dates), mold_schedule_frame=mold_schedule_frame)
        self.assertEqual(len(summary), 2)
        self.assertEqual(summary[0]["Heat Slot"], 1)
        self.assertEqual(summary[0]["Heat #"], 1)
        self.assertEqual(summary[0]["Anchor Alloy"], "LEW15")
        self.assertEqual(summary[0]["Total Weight (lbs)"], 1200.0)
        self.assertEqual(summary[0]["Total Molds"], 4.0)
        self.assertEqual(summary[0]["Max Mold Lead Days"], 0)
        self.assertEqual(summary[0]["Avg Mold Lead Days"], 0.0)
        self.assertEqual(summary[0]["Two Week Rule Status"], "VIOLATION")
        self.assertEqual(
            summary[0]["Job Breakout"],
            "5001-A | Due 08/04/2026 | Molds 2; 5001-B | Due 08/05/2026 | Molds 2",
        )
        self.assertEqual(summary[1]["Heat Status"], "Reserved")

    def test_build_heat_planner_rows_keeps_reserved_slot_blank_for_manual_fill(self):
        summary_rows = [
            {
                "Schedule Date": pd.Timestamp("2026-08-04").date(),
                "Weekday": "Tuesday",
                "Heat Slot": 6,
                "Heat #": "",
                "Heat Status": "Reserved",
                "Planning Priority": "",
                "Review Window": "",
                "Anchor Alloy": "",
                "Compatibility Group": "",
                "Earliest Due Date": "",
                "Latest Due Date": "",
                "Target Pour Date": "",
                "Pour Buffer Days": None,
                "Due Buffer Status": "",
                "Total Weight (lbs)": 0.0,
                "Total Molds": 0.0,
                "Rows in Heat": 0,
                "Jobs": "",
                "Extensions": "",
                "Planner Diagnostic": "",
            }
        ]

        planner_rows = build_heat_planner_rows(summary_rows)
        self.assertEqual(planner_rows[0]["Heat Slot"], 6)
        self.assertEqual(planner_rows[0]["Manual Alloy"], "")
        self.assertEqual(planner_rows[0]["Planner Notes"], "")

    def test_build_heat_detail_rows_flattens_planned_rows(self):
        melt_schedule = {
            1: {
                "rows": pd.DataFrame([
                    {
                        "Pour Schedule Day": 1,
                        "Heat #": 1,
                        "Global Heat #": 7,
                        "Planning Priority": "Highest Priority",
                        "Review Window": "Next 2 Weeks",
                        "Days Until Due": 3,
                        Columns.COL_DUE_DATE: "2026-08-07",
                        "Compatibility Group": "A216",
                        Columns.COL_ALLOY: "WCB",
                        Columns.COL_JOB_NUMBER: "7001",
                        "EXT": "A",
                        "Extension_Seq": 0,
                        "Molds for EXT": 2,
                        "Total Weight per EXT": 600,
                    }
                ])
            }
        }
        day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}

        detail_rows = build_heat_detail_rows(melt_schedule, day_dates)
        self.assertEqual(len(detail_rows), 1)
        self.assertEqual(detail_rows[0]["Heat #"], 1)
        self.assertEqual(detail_rows[0]["Global Heat #"], 7)
        self.assertEqual(detail_rows[0]["Job Number"], "7001")
        self.assertEqual(detail_rows[0]["Due Buffer Days"], 3)

    def test_export_heat_summary_writes_file(self):
        melt_schedule = {
            1: {
                "heat_summary": pd.DataFrame([
                    {
                        "Heat Slot": 1,
                        "Heat #": 1,
                        "Heat Status": "Planned",
                        "Planning Priority": "Highest Priority",
                        "Review Window": "Next 2 Weeks",
                        "Anchor Alloy": "LEW15",
                        "Compatibility Group": "A216",
                        "Earliest Due Date": pd.Timestamp("2026-08-04").date(),
                        "Latest Due Date": pd.Timestamp("2026-08-04").date(),
                        "Total Weight (lbs)": 600,
                        "Total Molds": 2,
                        "Rows in Heat": 1,
                        "Jobs": "5001",
                        "Extensions": "5001-A",
                    },
                    {
                        "Heat Slot": "",
                        "Heat #": 2,
                        "Heat Status": "Overflow",
                        "Planning Priority": "Highest Priority",
                        "Review Window": "Next 2 Weeks",
                        "Anchor Alloy": "LEW15",
                        "Compatibility Group": "A216",
                        "Earliest Due Date": pd.Timestamp("2026-08-04").date(),
                        "Latest Due Date": pd.Timestamp("2026-08-04").date(),
                        "Total Weight (lbs)": 400,
                        "Total Molds": 1,
                        "Rows in Heat": 1,
                        "Jobs": "5002",
                        "Extensions": "5002-B",
                    },
                    {
                        "Heat Slot": 6,
                        "Heat #": "",
                        "Heat Status": "Reserved",
                        "Planning Priority": "",
                        "Review Window": "",
                        "Anchor Alloy": "",
                        "Compatibility Group": "",
                        "Earliest Due Date": "",
                        "Latest Due Date": "",
                        "Total Weight (lbs)": 0,
                        "Total Molds": 0,
                        "Rows in Heat": 0,
                        "Jobs": "",
                        "Extensions": "",
                    },
                ]),
                "rows": pd.DataFrame([
                    {
                        Columns.COL_DUE_DATE: "2026-08-04",
                        "Customer Name": "Customer One",
                        "Part Number": "P1",
                        Columns.COL_ALLOY: "LEW15",
                        Columns.COL_JOB_NUMBER: "5001",
                        "EXT": "A",
                        "Molds for EXT": 2,
                        "Total Weight per EXT": 600,
                        "Heat #": 1,
                    },
                    {
                        Columns.COL_DUE_DATE: "2026-08-04",
                        "Customer Name": "Customer Two",
                        "Part Number": "P2",
                        Columns.COL_ALLOY: "LEW15",
                        Columns.COL_JOB_NUMBER: "5002",
                        "EXT": "B",
                        "Molds for EXT": 1,
                        "Total Weight per EXT": 400,
                        "Heat #": 2,
                    },
                ]),
            },
            2: {
                "heat_summary": pd.DataFrame([
                    {
                        "Heat Slot": 1,
                        "Heat #": 3,
                        "Heat Status": "Planned",
                        "Planning Priority": "Highest Priority",
                        "Review Window": "Next 2 Weeks",
                        "Anchor Alloy": "LEW20",
                        "Compatibility Group": "A217",
                        "Earliest Due Date": pd.Timestamp("2026-08-05").date(),
                        "Latest Due Date": pd.Timestamp("2026-08-05").date(),
                        "Total Weight (lbs)": 500,
                        "Total Molds": 1,
                        "Rows in Heat": 1,
                        "Jobs": "5003",
                        "Extensions": "5003-C",
                    }
                ]),
                "rows": pd.DataFrame([
                    {
                        Columns.COL_DUE_DATE: "2026-08-05",
                        "Customer Name": "Customer Three",
                        "Part Number": "P3",
                        Columns.COL_ALLOY: "LEW20",
                        Columns.COL_JOB_NUMBER: "5003",
                        "EXT": "C",
                        "Molds for EXT": 1,
                        "Total Weight per EXT": 500,
                        "Heat #": 3,
                    }
                ]),
            }
        }
        day_dates = {
            1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"},
            2: {"date": pd.Timestamp("2026-08-05"), "weekday": "Wednesday"},
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "heat_summary.xlsx"
            export_heat_summary(melt_schedule, day_dates, str(output_file))
            self.assertTrue(output_file.exists())

            wb = load_workbook(output_file)
            self.assertEqual(
                wb.sheetnames,
                ["Heat Summary", "Heat Planner", "Melt Mgmt Summary", "Melt Dept Schedule", "Melt Diagnostics"],
            )
            ws = wb["Heat Summary"]
            self.assertEqual(ws.cell(1, 1).value, "Heat Schedule")
            self.assertEqual(ws.cell(2, 1).value, "Date Created")
            self.assertEqual(ws.cell(2, 4).value, "Average Poured Lbs / Day")
            self.assertEqual(ws.cell(5, 1).value, "Heat Schedule")
            self.assertEqual(ws.cell(7, 1).value, "Schedule Date")
            self.assertEqual(ws.cell(7, 3).value, "Heat Slot")
            self.assertEqual(ws.cell(8, 3).value, 1)
            self.assertEqual(ws.cell(8, 8).value, "LEW15")
            self.assertEqual(ws.cell(7, 20).value, "Max Mold Lead Days")
            self.assertEqual(ws.cell(7, 21).value, "Avg Mold Lead Days")
            self.assertEqual(ws.cell(7, 22).value, "Two Week Rule Status")
            self.assertEqual(ws.cell(7, 23).value, "Two Week Rule Note")
            self.assertEqual(ws.cell(7, 24).value, "Job Breakout")
            self.assertEqual(ws.cell(8, 22).value, "VIOLATION")
            self.assertEqual(ws.cell(8, 22).fill.start_color.rgb, "00FFC7CE")
            self.assertEqual(
                ws.cell(8, 24).value,
                "5001-A | Due 08/04/2026 | Molds 2",
            )

            ws_planner = wb["Heat Planner"]
            self.assertEqual(ws_planner.cell(1, 19).value, "Max Mold Lead Days")
            self.assertEqual(ws_planner.cell(1, 20).value, "Avg Mold Lead Days")
            self.assertEqual(ws_planner.cell(1, 21).value, "Two Week Rule Status")
            self.assertEqual(ws_planner.cell(1, 22).value, "Two Week Rule Note")
            self.assertEqual(ws_planner.cell(1, 24).value, "Manual Alloy")
            self.assertEqual(ws_planner.cell(2, 21).value, "VIOLATION")
            self.assertEqual(ws_planner.cell(2, 21).fill.start_color.rgb, "00FFC7CE")
            self.assertEqual(ws_planner.cell(4, 3).value, 6)
            self.assertEqual(ws_planner.cell(4, 24).value, None)

            ws_melt_mgmt = wb["Melt Mgmt Summary"]
            self.assertEqual(ws_melt_mgmt.cell(1, 3).value, "Customer(s)")
            self.assertEqual(ws_melt_mgmt.cell(1, 4).value, "Poured Job IDs")
            self.assertEqual(ws_melt_mgmt.cell(1, 6).value, "Mold Date(s)")
            self.assertEqual(ws_melt_mgmt.cell(1, 12).value, "Planner Diagnostic")
            self.assertEqual(ws_melt_mgmt.cell(2, 4).value, "5001")
            self.assertEqual(ws_melt_mgmt.cell(7, 1).value, "Daily Melt Capacity")
            self.assertEqual(ws_melt_mgmt.cell(9, 1).value, "Day 1")
            self.assertEqual(ws_melt_mgmt.cell(9, 2).value, "Heats=2/5")
            self.assertEqual(ws_melt_mgmt.cell(9, 3).value, "Lbs=1000/10000")
            self.assertEqual(ws_melt_mgmt.cell(10, 1).value, "Day 2")
            self.assertEqual(ws_melt_mgmt.cell(10, 2).value, "Heats=1/5")
            self.assertEqual(ws_melt_mgmt.cell(10, 3).value, "Lbs=500/10000")

            ws_melt_dept = wb["Melt Dept Schedule"]
            self.assertEqual(ws_melt_dept.cell(5, 1).value, "Melt Schedule")
            self.assertEqual(ws_melt_dept.cell(7, 1).value, "Pour Date")
            self.assertEqual(ws_melt_dept.cell(7, 5).value, "Job Number")
            self.assertEqual(ws_melt_dept.cell(7, 6).value, "Molds on Floor")
            self.assertEqual(ws_melt_dept.cell(7, 7).value, "Pour Weight Required (lbs)")
            self.assertEqual(ws_melt_dept.cell(7, 8).value, "Part Number")
            self.assertEqual(ws_melt_dept.cell(8, 5).value, "5001")
            self.assertEqual(ws_melt_dept.cell(8, 6).value, 2)
            self.assertEqual(ws_melt_dept.cell(8, 7).value, 600)
            self.assertEqual(ws_melt_dept.cell(8, 8).value, "P1")
            self.assertEqual(ws_melt_dept.cell(8, 3).alignment.horizontal, None)
            self.assertEqual(ws_melt_dept.cell(8, 6).alignment.horizontal, "center")
            self.assertEqual(ws_melt_dept.cell(8, 7).alignment.horizontal, "center")
            self.assertEqual(ws_melt_dept.cell(8, 1).value.date().isoformat(), "2026-08-04")
            self.assertEqual(ws_melt_dept.cell(12, 1).value, "Melt Schedule")
            self.assertEqual(ws_melt_dept.cell(14, 1).value, "Pour Date")
            self.assertEqual(ws_melt_dept.cell(15, 1).value, datetime(2026, 8, 5, 0, 0))

            ws_melt_diag = wb["Melt Diagnostics"]
            self.assertEqual(ws_melt_diag.cell(1, 1).value, "Skipped Pour Day")
            self.assertEqual(ws_melt_diag.cell(1, 10).value, "Likely Cause")
            self.assertEqual(
                ws_melt_diag.cell(2, 1).value,
                "No skipped pour days detected in the current melt schedule range.",
            )

    def test_export_heat_summary_creates_missing_parent_directory(self):
        melt_schedule = {
            1: {
                "heat_summary": pd.DataFrame(),
                "rows": pd.DataFrame(),
            }
        }
        day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}

        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "nested" / "heat_summary.xlsx"
            export_heat_summary(melt_schedule, day_dates, str(output_file))
            self.assertTrue(output_file.exists())

    def test_export_combined_schedule_workbook_writes_expected_tabs(self):
        mold_rows = pd.DataFrame([
            {
                Columns.COL_DUE_DATE: "2026-08-04",
                "Customer Name": "Customer One",
                "Part Number": "P1",
                Columns.COL_JOB_NUMBER: "5001",
                "EXT": "A",
                Columns.COL_ALLOY: "LEW15",
                Columns.COL_CAST_TYPE: "L",
                "Quantity of Molds": 1,
                "Castings Per Mold": 1,
                "Quantity of Cores": 0,
                "Total Weight per EXT": 600,
                "Molds for EXT": 2,
                "Heat #": 1,
                "Pour Schedule Day": 1,
            }
        ])
        export_blocks = {
            1: {
                "date": pd.Timestamp("2026-08-04"),
                "weekday": "Tuesday",
                "rows": mold_rows,
                "weight_total": 600,
                "mold_total": 2,
            }
        }
        melt_schedule = {
            1: {
                "heat_summary": pd.DataFrame([
                    {
                        "Heat Slot": 1,
                        "Heat #": 1,
                        "Heat Status": "Planned",
                        "Planning Priority": "Highest Priority",
                        "Review Window": "Next 2 Weeks",
                        "Anchor Alloy": "LEW15",
                        "Compatibility Group": "A216",
                        "Earliest Due Date": pd.Timestamp("2026-08-04").date(),
                        "Latest Due Date": pd.Timestamp("2026-08-04").date(),
                        "Total Weight (lbs)": 600,
                        "Total Molds": 2,
                        "Rows in Heat": 1,
                        "Jobs": "5001",
                        "Extensions": "5001-A",
                    }
                ]),
                "rows": mold_rows,
            }
        }
        day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}
        job_shipping_rows = [
            {
                "Job Number": "5001",
                "Customer Name": "Customer One",
                "Schedule Status": "Scheduled",
                "Planned Molds": 2,
                "Scheduled Molds": 2,
                "Mold Day": 1,
                "Mold Date": pd.Timestamp("2026-08-04").date(),
                "Pour Day": 1,
                "Pour Date": pd.Timestamp("2026-08-04").date(),
                "Expected Ship Date": pd.Timestamp("2026-08-18").date(),
                "Due Date": pd.Timestamp("2026-08-20").date(),
                "Ship Buffer Days": 2,
                "On-Time": "YES",
            }
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "combined_schedule.xlsx"
            export_combined_schedule_workbook(
                export_blocks,
                melt_schedule,
                day_dates,
                output_file=str(output_file),
                job_shipping_rows=job_shipping_rows,
                mold_schedule_frame=mold_rows.assign(**{"Schedule Day": 1}),
                mold_day_dates=day_dates,
            )

            self.assertTrue(output_file.exists())
            wb = load_workbook(output_file)
            self.assertEqual(
                wb.sheetnames,
                [
                    "Overall Summary",
                    "Melt Schedule",
                    "Melt Diagnostics",
                    "Mold Schedule",
                    "Melt Summary",
                    "Mold Summary",
                ],
            )
            self.assertEqual(wb["Overall Summary"].cell(1, 1).value, "Job Number")
            self.assertEqual(wb["Melt Schedule"].cell(5, 1).value, "Melt Schedule")
            self.assertEqual(wb["Melt Schedule"].cell(7, 1).value, "Pour Date")
            self.assertEqual(wb["Melt Diagnostics"].cell(1, 1).value, "Skipped Pour Day")
            self.assertEqual(wb["Mold Schedule"].cell(1, 1).value, "Mold Schedule")
            self.assertEqual(wb["Melt Summary"].cell(1, 1).value, "Pour Date")
            self.assertEqual(wb["Mold Summary"].cell(1, 1).value, "Job Number")

    def test_build_heat_daily_totals_rows(self):
        summary_rows = [
            {
                "Schedule Date": pd.Timestamp("2026-08-04").date(),
                "Weekday": "Tuesday",
                "Heat Slot": 1,
                "Heat #": 1,
                "Heat Status": "Planned",
                "Anchor Alloy": "LEW15",
                "Due Buffer Status": "AT RISK",
                "Total Weight (lbs)": 600.0,
                "Total Molds": 2.0,
                "Rows in Heat": 1,
            },
            {
                "Schedule Date": pd.Timestamp("2026-08-04").date(),
                "Weekday": "Tuesday",
                "Heat Slot": "",
                "Heat #": 2,
                "Heat Status": "Overflow",
                "Anchor Alloy": "WCB",
                "Due Buffer Status": "ON TRACK",
                "Total Weight (lbs)": 900.0,
                "Total Molds": 3.0,
                "Rows in Heat": 2,
            },
            {
                "Schedule Date": pd.Timestamp("2026-08-04").date(),
                "Weekday": "Tuesday",
                "Heat Slot": 6,
                "Heat #": "",
                "Heat Status": "Reserved",
                "Anchor Alloy": "",
                "Due Buffer Status": "",
                "Total Weight (lbs)": 0.0,
                "Total Molds": 0.0,
                "Rows in Heat": 0,
            },
        ]

        daily_rows = build_heat_daily_totals_rows(summary_rows)
        self.assertEqual(len(daily_rows), 1)
        self.assertEqual(daily_rows[0]["Total Heats"], 2)
        self.assertEqual(daily_rows[0]["Planned Heats"], 1)
        self.assertEqual(daily_rows[0]["Overflow Heats"], 1)
        self.assertEqual(daily_rows[0]["At-Risk Heats"], 1)
        self.assertEqual(daily_rows[0]["Total Weight (lbs)"], 1500.0)
        self.assertEqual(daily_rows[0]["Total Molds"], 5.0)

    def test_export_heat_summary_wraps_save_failures(self):
        melt_schedule = {
            1: {
                "heat_summary": pd.DataFrame(),
            }
        }
        day_dates = {1: {"date": pd.Timestamp("2026-08-04"), "weekday": "Tuesday"}}

        with patch("fmes.scheduler_export.Workbook.save", side_effect=PermissionError("locked")):
            with self.assertRaises(RuntimeError) as context:
                export_heat_summary(melt_schedule, day_dates, "locked_heat.xlsx")

        self.assertIn("Failed while exporting heat summary", str(context.exception))


if __name__ == "__main__":
    unittest.main()