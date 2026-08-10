"""
Export logic for Foundry Management and Execution System (FMES).

Builds structured export blocks from daily schedules and writes two output
workbooks:
  Mold Schedule.xlsx  – per-day tables with job details, extension sizes, and heat numbers.
  Heat Summary.xlsx   – per-heat totals (Sheet 1) and per-day aggregate totals (Sheet 2).
"""

from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
import pandas as pd

from config import Columns


def _normalize_due_date(value):
    """Return a date object from value, an empty string for NaN, or the raw value if unparseable."""
    if pd.isna(value):
        return ""

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return value

    return parsed.date()


def build_daily_export_blocks(Daily_Schedules, Day_Dates):
    """
    Combine daily schedule DataFrames with their calendar dates into export blocks.

    Each block contains the subset of columns written to Excel plus pre-computed
    weight and mold totals used for the TOTALS row.

    Returns:
        dict mapping day number (int) -> {
            'date', 'weekday', 'rows' (DataFrame), 'weight_total', 'mold_total'
        }
    """
    try:
        export_blocks = {}

        for day, df in Daily_Schedules.items():
            df = df.copy()
            if "Heat #" not in df.columns:
                df["Heat #"] = ""

            weight_total = df["Total Weight per EXT"].fillna(0).sum()
            mold_total = df["Molds for EXT"].fillna(0).sum()

            export_blocks[day] = {
                "date": Day_Dates[day]["date"],
                "weekday": Day_Dates[day]["weekday"],
                "rows": df[
                    [
                        Columns.COL_DUE_DATE,
                        "Customer Name",
                        "Part Number",
                        Columns.COL_JOB_NUMBER,
                        "EXT",
                        Columns.COL_ALLOY,
                        Columns.COL_CAST_TYPE,
                        "Quantity of Molds",
                        "Castings Per Mold",
                        "Quantity of Cores",
                        "Total Weight per EXT",
                        "Molds for EXT",
                        "Heat #",
                    ]
                ].copy(),
                "weight_total": weight_total,
                "mold_total": mold_total,
            }

        return export_blocks
    except Exception as exc:
        raise RuntimeError("Failed while building export blocks") from exc


def print_export_blocks(export_blocks):
    """Print a formatted console preview of all export blocks."""
    try:
        for day in export_blocks:
            print("\n" + "=" * 50)
            print(
                f"Mold Schedule    "
                f"{export_blocks[day]['date'].strftime('%m/%d/%Y')}    "
                f"{export_blocks[day]['weekday']}"
            )
            print("=" * 50)
            print(export_blocks[day]["rows"])
            print(f"\nTOTAL WEIGHT: {export_blocks[day]['weight_total']}")
            print(f"TOTAL MOLDS: {export_blocks[day]['mold_total']}")
    except Exception as exc:
        raise RuntimeError("Failed while printing export blocks") from exc


def build_excel_rows(export_blocks):
    """
    Flatten export_blocks into a list of row lists for simple sequential writing.

    Each day contributes: a header row, a column-label row, data rows, a TOTALS
    row, and one blank spacer row.

    This is an alternate representation used for verification and lightweight
    consumers; Export_Mold_Schedule writes directly from structured blocks.

    Returns:
        list of lists – each inner list represents one Excel row.
    """
    excel_rows = []

    for day in sorted(export_blocks.keys()):
        block = export_blocks[day]
        excel_rows.append(["Mold Schedule", block["date"].strftime("%m/%d/%Y"), block["weekday"]])
        excel_rows.append([
            "Due Date",
            "Customer Name",
            "Part Number",
            "Job Number",
            "EXT",
            "Alloy",
            "Mold Type",
            "Quantity of Molds",
            "Castings Per Mold",
            "Cores Per Mold",
            "Total Weight per EXT",
            "# of Molds for EXT",
            "Heat #",
        ])

        for _, row in block["rows"].iterrows():
            excel_rows.append([
                _normalize_due_date(row.get(Columns.COL_DUE_DATE, "")),
                row.get("Customer Name", ""),
                row.get("Part Number", ""),
                row.get(Columns.COL_JOB_NUMBER, ""),
                row.get("EXT", ""),
                row.get(Columns.COL_ALLOY, ""),
                row.get(Columns.COL_CAST_TYPE, ""),
                row.get("Quantity of Molds", ""),
                row.get("Castings Per Mold", ""),
                row.get("Quantity of Cores", ""),
                row.get("Total Weight per EXT", ""),
                row.get("Molds for EXT", ""),
                row.get("Heat #", ""),
            ])

        excel_rows.append(["TOTALS", "", "", "", "", "", "", "", "", "", block["weight_total"], block["mold_total"], ""])
        excel_rows.append([])

    return excel_rows


def export_mold_schedule(Export_Blocks, output_file="Mold Schedule.xlsx"):
    """
    Write the mold schedule to an Excel workbook.

    Each production day occupies its own block of rows separated by two blank
    rows.  Column widths, bold headers, and thin borders are applied.

    Args:
        Export_Blocks: Output of Build_Daily_Export_Blocks.
        output_file:   Destination path for the workbook.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mold Schedule"
        current_row = 1

        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        bold = Font(bold=True)

        for day in sorted(Export_Blocks.keys()):
            block = Export_Blocks[day]

            ws.cell(current_row, 1, "Mold Schedule")
            ws.cell(current_row, 2, block["date"].strftime("%m/%d/%Y"))
            ws.cell(current_row, 4, block["weekday"])

            for col in range(1, 14):
                ws.cell(current_row, col).font = bold

            current_row += 2

            headers = [
                "Due Date",
                "Customer Name",
                "Part Number",
                "Job Number",
                "EXT",
                "Alloy",
                "Mold Type",
                "Quantity of Molds",
                "Castings Per Mold",
                "Cores Per Mold",
                "Total Weight per EXT",
                "# of Molds for EXT",
                "Heat #",
            ]

            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(current_row, col_num, header)
                cell.font = bold
                cell.border = thin

            current_row += 1

            for _, row in block["rows"].iterrows():
                values = [
                    _normalize_due_date(row.get(Columns.COL_DUE_DATE, "")),
                    row.get("Customer Name", ""),
                    row.get("Part Number", ""),
                    row.get(Columns.COL_JOB_NUMBER, ""),
                    row.get("EXT", ""),
                    row.get(Columns.COL_ALLOY, ""),
                    row.get(Columns.COL_CAST_TYPE, ""),
                    row.get("Quantity of Molds", ""),
                    row.get("Castings Per Mold", ""),
                    row.get("Quantity of Cores", ""),
                    row.get("Total Weight per EXT", ""),
                    row.get("Molds for EXT", ""),
                    row.get("Heat #", ""),
                ]

                for col_num, value in enumerate(values, start=1):
                    cell = ws.cell(current_row, col_num, value)
                    cell.border = thin
                    if col_num == 1 and value != "":
                        cell.number_format = "m/d/yyyy"

                current_row += 1

            ws.cell(current_row, 1, "TOTALS")
            ws.cell(current_row, 11, block["weight_total"])
            ws.cell(current_row, 12, block["mold_total"])
            ws.cell(current_row, 1).font = bold
            ws.cell(current_row, 11).font = bold
            ws.cell(current_row, 12).font = bold
            current_row += 3

        widths = {"A": 12, "B": 30, "C": 25, "D": 15, "E": 6, "F": 15, "G": 12, "H": 15, "I": 15, "J": 15, "K": 18, "L": 15}
        widths["M"] = 10

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        wb.save(output_file)
        print(f"Saved: {output_file}")
    except Exception as exc:
        raise RuntimeError(f"Failed while exporting mold schedule to {output_file}") from exc


def build_heat_summary_rows(export_blocks):
    """
    Aggregate export block rows into one summary row per heat per day.

    Returns:
        list of dicts with keys: Schedule Date, Weekday, Heat #, Alloy,
        Total Weight (lbs), Total Molds, Rows in Heat.
    """
    summary_rows = []

    for day in sorted(export_blocks.keys()):
        block = export_blocks[day]
        rows = block["rows"].copy()

        if rows.empty or "Heat #" not in rows.columns:
            continue

        rows["Total Weight per EXT"] = pd.to_numeric(
            rows["Total Weight per EXT"], errors="coerce"
        ).fillna(0)
        rows["Molds for EXT"] = pd.to_numeric(
            rows["Molds for EXT"], errors="coerce"
        ).fillna(0)

        grouped = rows.groupby("Heat #", dropna=False, sort=True)

        for heat_num, heat_df in grouped:
            if pd.isna(heat_num) or heat_num == "":
                continue

            alloy_values = [
                str(v)
                for v in heat_df[Columns.COL_ALLOY].dropna().unique().tolist()
                if str(v) != ""
            ]
            alloy = alloy_values[0] if alloy_values else ""

            summary_rows.append(
                {
                    "Schedule Date": block["date"].date() if hasattr(block["date"], "date") else block["date"],
                    "Weekday": block["weekday"],
                    "Heat #": int(heat_num),
                    "Alloy": alloy,
                    "Total Weight (lbs)": float(heat_df["Total Weight per EXT"].sum()),
                    "Total Molds": float(heat_df["Molds for EXT"].sum()),
                    "Rows in Heat": int(len(heat_df)),
                }
            )

    return summary_rows


def build_heat_daily_totals_rows(summary_rows):
    """
    Roll up heat summary rows to one row per production day.

    Returns:
        list of dicts with keys: Schedule Date, Weekday, Total Heats,
        Total Weight (lbs), Total Molds.
    """
    if not summary_rows:
        return []

    summary_df = pd.DataFrame(summary_rows)
    grouped = (
        summary_df
        .groupby(["Schedule Date", "Weekday"], dropna=False, sort=True)
        .agg(
            TotalHeats=("Heat #", "nunique"),
            TotalWeightLbs=("Total Weight (lbs)", "sum"),
            TotalMolds=("Total Molds", "sum"),
        )
        .reset_index()
    )

    daily_rows = []
    for _, row in grouped.iterrows():
        daily_rows.append(
            {
                "Schedule Date": row["Schedule Date"],
                "Weekday": row["Weekday"],
                "Total Heats": int(row["TotalHeats"]),
                "Total Weight (lbs)": float(row["TotalWeightLbs"]),
                "Total Molds": float(row["TotalMolds"]),
            }
        )

    return daily_rows


def export_heat_summary(export_blocks, output_file="Heat Summary.xlsx"):
    """
    Write the heat summary workbook with two sheets.

    Sheet 1 "Heat Summary"       – one row per heat per day.
    Sheet 2 "Daily Heat Totals"  – one row per day with aggregate counts.

    Args:
        export_blocks: Output of Build_Daily_Export_Blocks.
        output_file:   Destination path for the workbook.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Heat Summary"

        headers = [
            "Schedule Date",
            "Weekday",
            "Heat #",
            "Alloy",
            "Total Weight (lbs)",
            "Total Molds",
            "Rows in Heat",
        ]

        bold = Font(bold=True)
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_num, header)
            cell.font = bold
            cell.border = thin

        summary_rows = build_heat_summary_rows(export_blocks)
        current_row = 2

        for row in summary_rows:
            values = [
                row["Schedule Date"],
                row["Weekday"],
                row["Heat #"],
                row["Alloy"],
                row["Total Weight (lbs)"],
                row["Total Molds"],
                row["Rows in Heat"],
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(current_row, col_num, value)
                cell.border = thin
                if col_num == 1 and value != "":
                    cell.number_format = "m/d/yyyy"
                if col_num == 5:
                    cell.number_format = "#,##0.00"
                if col_num == 6:
                    cell.number_format = "#,##0"

            current_row += 1

        widths = {
            "A": 14,
            "B": 12,
            "C": 8,
            "D": 15,
            "E": 18,
            "F": 12,
            "G": 12,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws_daily = wb.create_sheet("Daily Heat Totals")
        daily_headers = [
            "Schedule Date",
            "Weekday",
            "Total Heats",
            "Total Weight (lbs)",
            "Total Molds",
        ]

        for col_num, header in enumerate(daily_headers, start=1):
            cell = ws_daily.cell(1, col_num, header)
            cell.font = bold
            cell.border = thin

        daily_rows = build_heat_daily_totals_rows(summary_rows)
        current_row = 2

        for row in daily_rows:
            values = [
                row["Schedule Date"],
                row["Weekday"],
                row["Total Heats"],
                row["Total Weight (lbs)"],
                row["Total Molds"],
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws_daily.cell(current_row, col_num, value)
                cell.border = thin
                if col_num == 1 and value != "":
                    cell.number_format = "m/d/yyyy"
                if col_num == 4:
                    cell.number_format = "#,##0.00"
                if col_num == 5:
                    cell.number_format = "#,##0"

            current_row += 1

        daily_widths = {
            "A": 14,
            "B": 12,
            "C": 12,
            "D": 18,
            "E": 12,
        }

        for col, width in daily_widths.items():
            ws_daily.column_dimensions[col].width = width

        wb.save(output_file)
        print(f"Saved: {output_file}")
    except Exception as exc:
        raise RuntimeError(f"Failed while exporting heat summary to {output_file}") from exc
