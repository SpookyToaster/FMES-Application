"""Workbook sync helpers for FMES Open Order Report updates."""

from datetime import datetime
from pathlib import Path
import shutil
import tempfile
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
import pandas as pd


XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XR_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
XR2_NS = "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"
XR3_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"


def to_plain_text(value):
    """Return plain-text cell value for workbook writes without touching formatting."""
    if pd.isna(value):
        return ""

    if isinstance(value, (datetime, pd.Timestamp)):
        return value.strftime("%Y-%m-%d")

    return str(value)


def _excel_column_letter(column_index):
    """Return Excel column letters for a 1-based column index."""
    letters = []
    while column_index > 0:
        column_index, remainder = divmod(column_index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _excel_column_index(column_letters):
    """Return 1-based numeric column index for Excel column letters."""
    value = 0
    for char in column_letters:
        value = (value * 26) + (ord(char.upper()) - 64)
    return value


def _sheet_cell_reference(row_index, column_index):
    """Return Excel-style cell reference like F2."""
    return f"{_excel_column_letter(column_index)}{row_index}"


def _cell_column_index(cell_reference):
    """Extract numeric column index from an Excel cell reference."""
    letters = "".join(char for char in cell_reference if char.isalpha())
    return _excel_column_index(letters)


def _resolve_sheet_archive_path(zip_file, sheet_name):
    """Resolve workbook sheet name to xl/worksheets/... archive path."""
    workbook_root = ET.fromstring(zip_file.read("xl/workbook.xml"))
    relationships_root = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))

    relationship_id = None
    for sheet in workbook_root.findall(f"{{{XML_NS}}}sheets/{{{XML_NS}}}sheet"):
        if sheet.get("name") == sheet_name:
            relationship_id = sheet.get(f"{{{REL_NS}}}id")
            break

    if not relationship_id:
        raise RuntimeError(f"Worksheet '{sheet_name}' was not found.")

    for relationship in relationships_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
        if relationship.get("Id") == relationship_id:
            target = relationship.get("Target", "")
            if target.startswith("/"):
                return target.lstrip("/")
            return f"xl/{target.lstrip('/')}"

    raise RuntimeError(f"Worksheet relationship for '{sheet_name}' was not found.")


def _find_or_create_row(sheet_data, row_index):
    """Return existing row element or insert a new one in sorted position."""
    row_tag = f"{{{XML_NS}}}row"

    for row in sheet_data.findall(row_tag):
        if int(row.get("r", "0")) == row_index:
            return row

    new_row = ET.Element(row_tag, {"r": str(row_index)})
    inserted = False
    for position, row in enumerate(sheet_data.findall(row_tag)):
        if int(row.get("r", "0")) > row_index:
            sheet_data.insert(position, new_row)
            inserted = True
            break

    if not inserted:
        sheet_data.append(new_row)

    return new_row


def _find_or_create_cell(row_element, row_index, column_index):
    """Return existing cell element or insert a new one in sorted position."""
    cell_ref = _sheet_cell_reference(row_index, column_index)
    cell_tag = f"{{{XML_NS}}}c"

    for cell in row_element.findall(cell_tag):
        if cell.get("r") == cell_ref:
            return cell

    new_cell = ET.Element(cell_tag, {"r": cell_ref})
    inserted = False
    for position, cell in enumerate(row_element.findall(cell_tag)):
        if _cell_column_index(cell.get("r", "A1")) > column_index:
            row_element.insert(position, new_cell)
            inserted = True
            break

    if not inserted:
        row_element.append(new_cell)

    return new_cell


def _set_cell_plain_text(cell, text):
    """Replace cell contents with inline plain text while preserving attributes/styles."""
    for child in list(cell):
        cell.remove(child)

    cell.set("t", "inlineStr")
    is_node = ET.SubElement(cell, f"{{{XML_NS}}}is")
    text_node = ET.SubElement(is_node, f"{{{XML_NS}}}t")
    text_node.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    text_node.text = text


def restore_ignorable_namespace_declarations(xml_bytes):
    """
    Restore namespace declarations referenced by Ignorable after ET serialization.

    ElementTree can rename/drop prefixes for unused namespaces, but worksheets may
    still include an Ignorable list that references those prefixes. Excel then
    repairs the sheet as malformed. This keeps the needed declarations present.
    """
    xml_text = xml_bytes.decode("utf-8")
    xml_decl_end = xml_text.find("?>")
    search_start = xml_decl_end + 2 if xml_decl_end >= 0 else 0
    root_tag_start = xml_text.find("<", search_start)
    if root_tag_start < 0:
        return xml_bytes

    root_tag_end = xml_text.find(">", root_tag_start)
    if root_tag_end <= root_tag_start:
        return xml_bytes

    root_open = xml_text[root_tag_start:root_tag_end]
    if "Ignorable=\"x14ac xr xr2 xr3\"" not in root_open:
        return xml_bytes

    missing_decls = []
    required = {
        "x14ac": X14AC_NS,
        "xr": XR_NS,
        "xr2": XR2_NS,
        "xr3": XR3_NS,
    }

    for prefix, uri in required.items():
        marker = f'xmlns:{prefix}="'
        if marker not in root_open:
            missing_decls.append(f' xmlns:{prefix}="{uri}"')

    if not missing_decls:
        return xml_bytes

    patched = xml_text[:root_tag_end] + "".join(missing_decls) + xml_text[root_tag_end:]
    return patched.encode("utf-8")


def export_worksheet_values(source_workbook_path, sheet_name, output_path):
    """Export raw worksheet values into a standalone workbook."""
    source_wb = load_workbook(source_workbook_path, data_only=True)
    try:
        if sheet_name not in source_wb.sheetnames:
            raise RuntimeError(f"Worksheet '{sheet_name}' was not found.")

        source_ws = source_wb[sheet_name]

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = sheet_name

        for row in source_ws.iter_rows(
            min_row=1,
            max_row=source_ws.max_row,
            min_col=1,
            max_col=source_ws.max_column,
            values_only=True,
        ):
            new_ws.append(list(row))

        new_wb.save(output_path)
    finally:
        source_wb.close()


def write_sql_data_to_oor(source_workbook_path, sql_rows, sql_main_export_columns, sheet_name="OOR"):
    """Overwrite OOR F:V values by editing sheet XML directly to preserve workbook metadata."""
    start_row = 2
    start_col = 6
    end_col = 22

    with zipfile.ZipFile(source_workbook_path, "r") as source_zip:
        sheet_archive_path = _resolve_sheet_archive_path(source_zip, sheet_name)
        sheet_root = ET.fromstring(source_zip.read(sheet_archive_path))
        sheet_data = sheet_root.find(f"{{{XML_NS}}}sheetData")

        if sheet_data is None:
            raise RuntimeError(f"Worksheet '{sheet_name}' does not contain sheet data.")

        max_existing_row = 0
        for row in sheet_data.findall(f"{{{XML_NS}}}row"):
            max_existing_row = max(max_existing_row, int(row.get("r", "0")))

        if max_existing_row >= start_row:
            for row_idx in range(start_row, max_existing_row + 1):
                row_element = _find_or_create_row(sheet_data, row_idx)
                for col_idx in range(start_col, end_col + 1):
                    cell = _find_or_create_cell(row_element, row_idx, col_idx)
                    _set_cell_plain_text(cell, "")

        for offset, sql_row in enumerate(sql_rows):
            row_idx = start_row + offset
            row_element = _find_or_create_row(sheet_data, row_idx)
            for col_offset, col_name in enumerate(sql_main_export_columns):
                col_idx = start_col + col_offset
                cell = _find_or_create_cell(row_element, row_idx, col_idx)
                _set_cell_plain_text(cell, to_plain_text(sql_row.get(col_name, "")))

        updated_sheet_xml = ET.tostring(
            sheet_root,
            encoding="utf-8",
            xml_declaration=True,
        )
        updated_sheet_xml = restore_ignorable_namespace_declarations(updated_sheet_xml)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_path = temp_file.name

        try:
            with zipfile.ZipFile(temp_path, "w") as target_zip:
                for zip_info in source_zip.infolist():
                    if zip_info.filename == sheet_archive_path:
                        target_zip.writestr(zip_info, updated_sheet_xml)
                    else:
                        target_zip.writestr(zip_info, source_zip.read(zip_info.filename))
            shutil.move(temp_path, source_workbook_path)
        finally:
            if Path(temp_path).exists():
                Path(temp_path).unlink(missing_ok=True)


def save_sql_snapshot(sql_rows, sql_main_export_columns, output_path):
    """Save SQL rows to a standalone snapshot workbook for day-to-day diffing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SQL Snapshot"

    ws.append(sql_main_export_columns)
    for row in sql_rows:
        ws.append([to_plain_text(row.get(col, "")) for col in sql_main_export_columns])

    wb.save(output_path)