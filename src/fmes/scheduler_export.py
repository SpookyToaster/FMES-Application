"""
Export logic for Foundry Management and Execution System (FMES).

Builds structured export blocks from daily schedules and writes two output
workbooks:
    Mold Schedule.xlsx  – per-day tables with job details, extension sizes, and heat numbers.
    Heat Summary.xlsx   – melt-plan summary, daily totals, planner worksheet,
                          and row-level heat-plan detail.
"""

from copy import copy
from pathlib import Path
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Font, Side
from openpyxl.styles import PatternFill
import pandas as pd

from .config import Columns, DailyMoldLimits
from .melt_planning import DAILY_WEIGHT_TARGET_LBS, MAX_PLANNED_HEATS_PER_DAY


def _ensure_output_parent(output_file):
    """Create the destination parent directory when it does not already exist."""
    output_path = Path(output_file)
    if output_path.parent and not output_path.parent.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)


def _normalize_due_date(value):
    """Return a date object from value, an empty string for NaN, or the raw value if unparseable."""
    if pd.isna(value):
        return ""

    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return value

    return parsed.date()


def _normalize_date_value(value):
    """Return normalized pandas Timestamp date value or NaT."""
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return pd.NaT
    return pd.Timestamp(parsed).normalize()


def _target_pour_date(due_date, target_days=14):
    """Return the target pour date that satisfies the minimum lead-time goal."""
    if pd.isna(due_date):
        return pd.NaT
    return due_date - pd.Timedelta(days=target_days)


def _planner_risk_for_buffer(buffer_days):
    """Return planner-facing risk status and diagnostic text for pour buffer."""
    if buffer_days is None:
        return "UNKNOWN", "No due date available"
    if buffer_days < 14:
        return "AT RISK", "Pour is less than 14 days before due"
    if buffer_days < 21:
        return "WATCH", "Pour is within a narrow due-date buffer window"
    return "ON TRACK", "Pour meets the minimum 14-day due buffer"


def _apply_11x17_portrait_layout(ws):
    """Apply print settings optimized for 11x17 portrait output."""
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _apply_letter_portrait_layout(ws):
    """Apply print settings optimized for letter portrait output."""
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def _apply_fill_to_row(ws, row_num, column_count, fill):
    """Apply a fill style across a contiguous row segment."""
    for col_num in range(1, column_count + 1):
        ws.cell(row_num, col_num).fill = fill


def _insert_day_break_row(ws, row_num, column_count, fill=None, height=8):
    """Insert a blank visual spacer row between schedule days."""
    ws.row_dimensions[row_num].height = height
    if fill is not None:
        _apply_fill_to_row(ws, row_num, column_count, fill)


REPORT_TITLE_FILL = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
REPORT_SUBTITLE_FILL = PatternFill(fill_type="solid", start_color="D9E2F3", end_color="D9E2F3")
REPORT_HEADER_FILL = PatternFill(fill_type="solid", start_color="D9E2F3", end_color="D9E2F3")
REPORT_TOTAL_FILL = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")
REPORT_SEPARATOR_FILL = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
MOLD_LEAD_DIAGNOSTIC_WINDOW_DAYS = 5


def _write_schedule_banner(ws, title, created_on, metric_label, metric_value, last_column):
    """Write a consistent top-of-sheet banner for schedule workbooks."""
    title_font = Font(bold=True, color="FFFFFF", size=14)
    label_font = Font(bold=True)
    value_font = Font(bold=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = ws.cell(1, 1, title)
    title_cell.font = title_font
    title_cell.fill = REPORT_TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    created_label = ws.cell(2, 1, "Date Created")
    created_label.font = label_font
    created_label.fill = REPORT_SUBTITLE_FILL
    created_value = ws.cell(2, 2, created_on.strftime("%m/%d/%Y"))
    created_value.font = value_font
    created_value.fill = REPORT_SUBTITLE_FILL

    metric_label_cell = ws.cell(2, 4, metric_label)
    metric_label_cell.font = label_font
    metric_label_cell.fill = REPORT_SUBTITLE_FILL
    metric_value_cell = ws.cell(2, 5, metric_value)
    metric_value_cell.font = value_font
    metric_value_cell.fill = REPORT_SUBTITLE_FILL

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8


def _apply_schedule_day_header(ws, row_num, column_count, fill=REPORT_HEADER_FILL):
    """Apply consistent header styling to a day block header row."""
    _apply_fill_to_row(ws, row_num, column_count, fill)
    for col_num in range(1, column_count + 1):
        cell = ws.cell(row_num, col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_schedule_total_row(ws, row_num, columns):
    """Apply consistent styling to a totals row."""
    for col_num in columns:
        ws.cell(row_num, col_num).fill = REPORT_TOTAL_FILL
        ws.cell(row_num, col_num).font = Font(bold=True)


def _apply_schedule_data_row(ws, row_num, column_count):
    """Apply a consistent border/fill baseline to schedule data rows."""
    for col_num in range(1, column_count + 1):
        ws.cell(row_num, col_num).fill = REPORT_SEPARATOR_FILL


def _build_mold_capacity_rows(export_blocks):
    """Summarize daily line/floor mold usage against configured day limits."""
    capacity_rows = []

    for day in sorted(export_blocks.keys()):
        rows = export_blocks[day].get("rows", pd.DataFrame())
        if rows.empty:
            line_molds = 0
            floor_molds = 0
        else:
            cast_types = rows.get(Columns.COL_CAST_TYPE, pd.Series("", index=rows.index)).fillna("").astype(str).str.strip().str.upper()
            mold_counts = pd.to_numeric(rows.get("Molds for EXT", pd.Series(0, index=rows.index)), errors="coerce").fillna(0)
            floor_mask = cast_types == "F"
            floor_molds = int(mold_counts[floor_mask].sum())
            line_molds = int(mold_counts[~floor_mask].sum())

        capacity_rows.append(
            {
                "Day": f"Day {day}",
                "Line Capacity": f"L={line_molds}/{DailyMoldLimits.MAX_L_MOLDS_PER_DAY}",
                "Floor Capacity": f"F={floor_molds}/{DailyMoldLimits.MAX_F_MOLDS_PER_DAY}",
            }
        )

    return capacity_rows


def _build_melt_capacity_rows(melt_schedule):
    """Summarize daily melt utilization against heat-count and weight targets."""
    capacity_rows = []

    for day in sorted(melt_schedule.keys()):
        rows = melt_schedule[day].get("rows", pd.DataFrame())
        if rows.empty:
            heat_count = 0
            total_weight = 0
        else:
            heat_numbers = rows.get("Heat #", pd.Series(dtype="object"))
            heat_numbers = heat_numbers[(heat_numbers != "") & heat_numbers.notna()]
            heat_count = int(pd.Series(heat_numbers).nunique())
            total_weight = int(round(pd.to_numeric(rows.get("Total Weight per EXT", pd.Series(0, index=rows.index)), errors="coerce").fillna(0).sum()))

        capacity_rows.append(
            {
                "Day": f"Day {day}",
                "Heat Capacity": f"Heats={heat_count}/{MAX_PLANNED_HEATS_PER_DAY}",
                "Weight Capacity": f"Lbs={total_weight}/{DAILY_WEIGHT_TARGET_LBS}",
            }
        )

    return capacity_rows


def _add_business_days(start_date, offset_days):
    """Return a weekday-only date shifted by offset_days from start_date."""
    if pd.isna(start_date):
        return pd.NaT

    current = pd.Timestamp(start_date).normalize()
    if offset_days == 0:
        return current

    step = 1 if offset_days > 0 else -1
    remaining = abs(int(offset_days))
    while remaining > 0:
        current += pd.Timedelta(days=step)
        while current.weekday() > 4:
            current += pd.Timedelta(days=step)
        remaining -= 1
    return current


def _resolve_schedule_day_date(day_value, day_dates):
    """Resolve a schedule day to a calendar date, even if day is not mapped."""
    if not day_dates:
        return pd.NaT

    try:
        day_int = int(day_value)
    except (TypeError, ValueError):
        return pd.NaT

    if day_int in day_dates:
        return _normalize_date_value(day_dates.get(day_int, {}).get("date", pd.NaT))

    known_days = sorted(int(day) for day in day_dates.keys())
    if not known_days:
        return pd.NaT

    anchor_day = known_days[0]
    anchor_date = _normalize_date_value(day_dates.get(anchor_day, {}).get("date", pd.NaT))
    if pd.isna(anchor_date):
        return pd.NaT

    return _add_business_days(anchor_date, day_int - anchor_day)


def _build_melt_gap_diagnostics(melt_schedule, day_dates, mold_schedule_frame):
    """Summarize skipped pour days and likely causes using mold-window capacity evidence."""
    if not melt_schedule:
        return []

    pour_days = sorted(int(day) for day in melt_schedule.keys())
    if not pour_days:
        return []

    present_days = set(pour_days)

    day_usage = {}
    if mold_schedule_frame is not None and not mold_schedule_frame.empty and "Schedule Day" in mold_schedule_frame.columns:
        mold_frame = mold_schedule_frame.copy()
        cast_types = mold_frame.get(Columns.COL_CAST_TYPE, pd.Series("", index=mold_frame.index)).fillna("").astype(str).str.strip().str.upper()
        mold_counts = pd.to_numeric(mold_frame.get("Molds for EXT", pd.Series(0, index=mold_frame.index)), errors="coerce").fillna(0)
        schedule_days = pd.to_numeric(mold_frame.get("Schedule Day", pd.Series(dtype="float64")), errors="coerce")

        for day_value, cast_type, mold_count in zip(
            schedule_days.tolist(),
            cast_types.tolist(),
            mold_counts.tolist(),
        ):
            if pd.isna(day_value):
                continue
            day_int = int(day_value)
            usage = day_usage.setdefault(day_int, {"L": 0, "F": 0})
            molds = int(round(float(mold_count)))
            bucket = "F" if cast_type == "F" else "L"
            usage[bucket] += max(molds, 0)

    heat_moves = pd.DataFrame()
    if (
        mold_schedule_frame is not None
        and not mold_schedule_frame.empty
        and "Original Pour Schedule Day" in mold_schedule_frame.columns
        and "Pour Schedule Day" in mold_schedule_frame.columns
        and "Heat #" in mold_schedule_frame.columns
    ):
        heat_moves = mold_schedule_frame[["Original Pour Schedule Day", "Pour Schedule Day", "Heat #"]].copy()
        heat_moves["Original Pour Schedule Day"] = pd.to_numeric(
            heat_moves["Original Pour Schedule Day"],
            errors="coerce",
        )
        heat_moves["Pour Schedule Day"] = pd.to_numeric(
            heat_moves["Pour Schedule Day"],
            errors="coerce",
        )
        heat_moves = heat_moves.dropna(subset=["Original Pour Schedule Day", "Pour Schedule Day", "Heat #"]).copy()
        if not heat_moves.empty:
            heat_moves["Original Pour Schedule Day"] = heat_moves["Original Pour Schedule Day"].astype(int)
            heat_moves["Pour Schedule Day"] = heat_moves["Pour Schedule Day"].astype(int)
            heat_moves = heat_moves.drop_duplicates()

    diagnostic_rows = []
    for day in range(min(pour_days), max(pour_days) + 1):
        if day in present_days:
            continue

        window_start = max(1, day - MOLD_LEAD_DIAGNOSTIC_WINDOW_DAYS)
        window_days = list(range(window_start, day))
        line_capacity = len(window_days) * DailyMoldLimits.MAX_L_MOLDS_PER_DAY
        floor_capacity = len(window_days) * DailyMoldLimits.MAX_F_MOLDS_PER_DAY
        line_used = sum(day_usage.get(window_day, {}).get("L", 0) for window_day in window_days)
        floor_used = sum(day_usage.get(window_day, {}).get("F", 0) for window_day in window_days)
        line_free = max(line_capacity - line_used, 0)
        floor_free = max(floor_capacity - floor_used, 0)

        originally_planned = 0
        pushed_later = 0
        pushed_to_days = ""
        if not heat_moves.empty:
            original_mask = heat_moves["Original Pour Schedule Day"] == day
            originally_planned = int(original_mask.sum())
            pushed_mask = original_mask & (heat_moves["Pour Schedule Day"] > day)
            pushed_later = int(pushed_mask.sum())
            pushed_days = sorted(int(value) for value in heat_moves.loc[pushed_mask, "Pour Schedule Day"].unique())
            pushed_to_days = ", ".join(str(value) for value in pushed_days)

        if originally_planned <= 0:
            likely_cause = "No heats remained targeted to this day after sequencing and prioritization."
        elif pushed_later <= 0:
            likely_cause = "No final pours landed on this day after melt-plan rebuild."
        elif line_free <= 0 and floor_free <= 0:
            likely_cause = "Original heats were pushed later because no pre-pour line or floor mold capacity remained in the 5-day window."
        elif line_free <= 0:
            likely_cause = "Original heats were pushed later because line mold capacity was fully consumed in the 5-day window."
        elif floor_free <= 0:
            likely_cause = "Original heats were pushed later because floor mold capacity was fully consumed in the 5-day window."
        else:
            likely_cause = "Original heats were pushed later to keep heat sequence and mold constraints feasible."

        resolved_date = _resolve_schedule_day_date(day, day_dates)
        diagnostic_rows.append(
            {
                "Skipped Pour Day": day,
                "Skipped Pour Date": _normalize_due_date(resolved_date) if not pd.isna(resolved_date) else "",
                "Weekday": pd.Timestamp(resolved_date).strftime("%A") if not pd.isna(resolved_date) else "",
                "Original Heats On Day": originally_planned,
                "Heats Pushed Later": pushed_later,
                "Pushed To Pour Days": pushed_to_days,
                "Mold Window Days": f"{window_start}-{day - 1}" if window_days else "",
                "Window Line Usage": f"{line_used}/{line_capacity}",
                "Window Floor Usage": f"{floor_used}/{floor_capacity}",
                "Likely Cause": likely_cause,
            }
        )

    return diagnostic_rows


def _write_capacity_block(ws, start_row, title, headers, rows, bold, thin):
    """Write a compact capacity summary block and return the next free row."""
    ws.cell(start_row, 1, title).font = bold

    header_row = start_row + 1
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_num, header)
        cell.font = bold
        cell.border = thin

    data_row = header_row + 1
    for row in rows:
        values = [row.get(header, "") for header in headers]
        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(data_row, col_num, value)
            cell.border = thin
        data_row += 1

    return data_row


def _clone_sheet_to_workbook(source_ws, target_wb, target_title):
    """Clone values, styles, and print layout from one workbook sheet to another."""
    target_ws = target_wb.create_sheet(target_title)

    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.font is not None:
                target_cell.font = copy(cell.font)
            if cell.fill is not None:
                target_cell.fill = copy(cell.fill)
            if cell.border is not None:
                target_cell.border = copy(cell.border)
            if cell.alignment is not None:
                target_cell.alignment = copy(cell.alignment)
            if cell.protection is not None:
                target_cell.protection = copy(cell.protection)
            if cell.number_format:
                target_cell.number_format = cell.number_format
            if cell.comment is not None:
                target_cell.comment = copy(cell.comment)
            if cell.hyperlink:
                target_cell._hyperlink = copy(cell.hyperlink)

    for key, dimension in source_ws.column_dimensions.items():
        target_ws.column_dimensions[key].width = dimension.width
        target_ws.column_dimensions[key].hidden = dimension.hidden

    for key, dimension in source_ws.row_dimensions.items():
        target_ws.row_dimensions[key].height = dimension.height
        target_ws.row_dimensions[key].hidden = dimension.hidden

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.auto_filter.ref = source_ws.auto_filter.ref
    target_ws.page_setup.orientation = source_ws.page_setup.orientation
    target_ws.page_setup.paperSize = source_ws.page_setup.paperSize
    target_ws.page_setup.fitToWidth = source_ws.page_setup.fitToWidth
    target_ws.page_setup.fitToHeight = source_ws.page_setup.fitToHeight
    target_ws.page_margins.left = source_ws.page_margins.left
    target_ws.page_margins.right = source_ws.page_margins.right
    target_ws.page_margins.top = source_ws.page_margins.top
    target_ws.page_margins.bottom = source_ws.page_margins.bottom
    target_ws.print_options.horizontalCentered = source_ws.print_options.horizontalCentered
    target_ws.print_options.verticalCentered = source_ws.print_options.verticalCentered
    target_ws.print_area = source_ws.print_area
    target_ws.print_title_cols = source_ws.print_title_cols
    target_ws.print_title_rows = source_ws.print_title_rows

    return target_ws


def export_combined_schedule_workbook(
    export_blocks,
    melt_schedule,
    pour_day_dates,
    output_file="Production Schedule Summary.xlsx",
    job_shipping_rows=None,
    mold_schedule_frame=None,
    mold_day_dates=None,
):
    """Export one combined workbook with management summaries and department schedules."""
    try:
        _ensure_output_parent(output_file)

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_dir_path = Path(temp_dir)
            mold_path = temp_dir_path / "mold_tmp.xlsx"
            heat_path = temp_dir_path / "heat_tmp.xlsx"

            export_mold_schedule(
                export_blocks,
                str(mold_path),
                job_shipping_rows=job_shipping_rows or [],
            )
            export_heat_summary(
                melt_schedule,
                pour_day_dates,
                str(heat_path),
                mold_schedule_frame=mold_schedule_frame,
                mold_day_dates=mold_day_dates,
            )

            from openpyxl import load_workbook

            source_mold_wb = load_workbook(str(mold_path))
            source_heat_wb = load_workbook(str(heat_path))

            combined_wb = Workbook()
            combined_wb.remove(combined_wb.active)

            sheet_plan = [
                (source_mold_wb, "Overall Job Status", "Overall Summary"),
                (source_heat_wb, "Melt Mgmt Summary", "Melt Summary"),
                (source_heat_wb, "Melt Dept Schedule", "Melt Schedule"),
                (source_heat_wb, "Melt Diagnostics", "Melt Diagnostics"),
                (source_mold_wb, "Mold Schedule", "Mold Schedule"),
            ]

            for source_wb, source_name, target_name in sheet_plan:
                if source_name not in source_wb.sheetnames:
                    raise RuntimeError(f"Missing expected sheet '{source_name}' while building combined workbook")
                _clone_sheet_to_workbook(source_wb[source_name], combined_wb, target_name)

            combined_wb.save(output_file)
            print(f"Saved: {output_file}")
    except Exception as exc:
        raise RuntimeError(f"Failed while exporting combined schedule workbook to {output_file}") from exc


def build_job_shipping_report_rows(
    schedule_data_frame,
    mold_schedule_frame,
    mold_day_dates,
    pour_day_dates,
    cleaning_days=14,
):
    """Build job-level shipping outlook rows from backfilled mold/pour assignments."""
    if schedule_data_frame is None or schedule_data_frame.empty:
        return []

    base = schedule_data_frame.copy()
    base[Columns.COL_DUE_DATE] = base[Columns.COL_DUE_DATE].apply(_normalize_date_value)

    planned_molds = (
        base.groupby(Columns.COL_JOB_NUMBER, dropna=False)["Molds for EXT"]
        .sum()
        .rename("Planned Molds")
    )
    due_dates = (
        base.groupby(Columns.COL_JOB_NUMBER, dropna=False)[Columns.COL_DUE_DATE]
        .min()
        .rename("Due Date")
    )
    if "Customer Name" in base.columns:
        customers = (
            base.groupby(Columns.COL_JOB_NUMBER, dropna=False)["Customer Name"]
            .agg(lambda values: next((str(v).strip() for v in values if str(v).strip()), ""))
            .rename("Customer Name")
        )
    else:
        customers = (
            base.groupby(Columns.COL_JOB_NUMBER, dropna=False)
            .size()
            .rename("Customer Name")
            .map(lambda _: "")
        )

    job_summary = pd.concat([planned_molds, due_dates, customers], axis=1).reset_index()

    if mold_schedule_frame is None or mold_schedule_frame.empty:
        job_summary["Scheduled Molds"] = 0
        job_summary["Mold Day"] = pd.NA
        job_summary["Pour Day"] = pd.NA
    else:
        assigned = mold_schedule_frame.copy()
        assigned_molds = (
            assigned.groupby(Columns.COL_JOB_NUMBER, dropna=False)["Molds for EXT"]
            .sum()
            .rename("Scheduled Molds")
        )
        mold_day = (
            assigned.groupby(Columns.COL_JOB_NUMBER, dropna=False)["Schedule Day"]
            .min()
            .rename("Mold Day")
        )
        pour_day = (
            assigned.groupby(Columns.COL_JOB_NUMBER, dropna=False)["Pour Schedule Day"]
            .max()
            .rename("Pour Day")
        )
        assigned_summary = pd.concat([assigned_molds, mold_day, pour_day], axis=1).reset_index()
        job_summary = job_summary.merge(
            assigned_summary,
            on=Columns.COL_JOB_NUMBER,
            how="left",
        )
        job_summary["Scheduled Molds"] = job_summary["Scheduled Molds"].fillna(0)

    def _resolve_day_date(day_value, day_map):
        try:
            day_key = int(day_value)
        except (TypeError, ValueError):
            return pd.NaT
        return _normalize_date_value(day_map.get(day_key, {}).get("date", pd.NaT))

    job_summary["Mold Date"] = job_summary["Mold Day"].apply(lambda value: _resolve_day_date(value, mold_day_dates))
    job_summary["Pour Date"] = job_summary["Pour Day"].apply(lambda value: _resolve_day_date(value, pour_day_dates))
    job_summary["Expected Ship Date"] = job_summary["Pour Date"].apply(
        lambda pour_date: pour_date + pd.Timedelta(days=cleaning_days) if not pd.isna(pour_date) else pd.NaT
    )

    job_summary["Ship Buffer Days"] = [
        int((due_date - expected_ship).days)
        if not pd.isna(due_date) and not pd.isna(expected_ship)
        else None
        for due_date, expected_ship in zip(job_summary["Due Date"], job_summary["Expected Ship Date"])
    ]

    def _schedule_status(row):
        planned = float(row.get("Planned Molds", 0) or 0)
        scheduled = float(row.get("Scheduled Molds", 0) or 0)
        if scheduled <= 0:
            return "Not Yet Scheduled"
        if scheduled + 1e-9 < planned:
            return "Partially Scheduled"
        return "Scheduled"

    def _on_time_status(row):
        status = row.get("Schedule Status", "")
        buffer_days = row.get("Ship Buffer Days", None)
        if status == "Not Yet Scheduled":
            return "NOT SCHEDULED"
        if buffer_days is None:
            return "UNKNOWN"
        return "YES" if buffer_days >= 0 else "NO"

    job_summary["Schedule Status"] = job_summary.apply(_schedule_status, axis=1)
    job_summary["On-Time"] = job_summary.apply(_on_time_status, axis=1)

    job_summary = job_summary.sort_values(
        by=["Schedule Status", Columns.COL_DUE_DATE, Columns.COL_JOB_NUMBER],
        ascending=[True, True, True],
        na_position="last",
    )

    rows = []
    for _, row in job_summary.iterrows():
        rows.append(
            {
                "Job Number": row.get(Columns.COL_JOB_NUMBER, ""),
                "Schedule Status": row.get("Schedule Status", ""),
                "Customer Name": row.get("Customer Name", ""),
                "Planned Molds": int(row.get("Planned Molds", 0) or 0),
                "Scheduled Molds": int(row.get("Scheduled Molds", 0) or 0),
                "Mold Day": int(row["Mold Day"]) if pd.notna(row.get("Mold Day", pd.NA)) else "",
                "Mold Date": _normalize_due_date(row.get("Mold Date", "")),
                "Pour Day": int(row["Pour Day"]) if pd.notna(row.get("Pour Day", pd.NA)) else "",
                "Pour Date": _normalize_due_date(row.get("Pour Date", "")),
                "Expected Ship Date": _normalize_due_date(row.get("Expected Ship Date", "")),
                "Due Date": _normalize_due_date(row.get("Due Date", "")),
                "Ship Buffer Days": row.get("Ship Buffer Days", ""),
                "On-Time": row.get("On-Time", ""),
            }
        )

    return rows


def build_daily_export_blocks(Daily_Schedules, Day_Dates, pour_day_dates=None):
    """
    Combine daily schedule DataFrames with their calendar dates into export blocks.

    Each block contains the subset of columns written to Excel plus pre-computed
    weight and mold totals used for the TOTALS row.

    Returns:
        dict mapping day number (int) -> {
            'date', 'weekday', 'rows' (DataFrame), 'weight_total', 'mold_total'
        }
    """
    try:
        export_blocks = {}

        for day, df in Daily_Schedules.items():
            df = df.copy()
            if "Heat #" not in df.columns:
                df["Heat #"] = ""

            for required_col in [Columns.COL_ALLOY, Columns.COL_DUE_DATE, Columns.COL_JOB_NUMBER]:
                if required_col not in df.columns:
                    df[required_col] = ""

            df = df.sort_values(
                by=[Columns.COL_ALLOY, Columns.COL_DUE_DATE, Columns.COL_JOB_NUMBER],
                ascending=[True, True, True],
                na_position="last",
            ).reset_index(drop=True)

            mold_day = Day_Dates[day]

            def _resolve_pour_day_value(row):
                pour_day = row.get("Pour Schedule Day", "")
                if pd.isna(pour_day):
                    return ""
                try:
                    return int(pour_day)
                except (TypeError, ValueError):
                    return ""

            df["Pour Schedule Day"] = df.apply(_resolve_pour_day_value, axis=1)
            if pour_day_dates is None:
                df["Pour Date"] = ""
                df["Pour Weekday"] = ""
            else:
                df["Pour Date"] = df["Pour Schedule Day"].apply(
                    lambda pour_day: pour_day_dates.get(pour_day, {}).get("date", "") if pour_day != "" else ""
                )
                df["Pour Weekday"] = df["Pour Schedule Day"].apply(
                    lambda pour_day: pour_day_dates.get(pour_day, {}).get("weekday", "") if pour_day != "" else ""
                )

            df["Mold Date"] = mold_day["date"]
            df["Mold Weekday"] = mold_day["weekday"]

            due_dates = df[Columns.COL_DUE_DATE].apply(_normalize_date_value)
            pour_dates = df["Pour Date"].apply(_normalize_date_value)
            df["Pour Buffer Days"] = [
                int((due_date - pour_date).days)
                if not pd.isna(due_date) and not pd.isna(pour_date)
                else None
                for due_date, pour_date in zip(due_dates, pour_dates)
            ]

            risk_status = df["Pour Buffer Days"].apply(_planner_risk_for_buffer)
            df["Due Buffer Status"] = risk_status.apply(lambda pair: pair[0])
            df["Planner Diagnostic"] = risk_status.apply(lambda pair: pair[1])

            weight_total = df["Total Weight per EXT"].fillna(0).sum()
            mold_total = df["Molds for EXT"].fillna(0).sum()

            export_blocks[day] = {
                "date": Day_Dates[day]["date"],
                "weekday": Day_Dates[day]["weekday"],
                "rows": df[
                    [
                        Columns.COL_DUE_DATE,
                        "Customer Name",
                        "Part Number",
                        Columns.COL_JOB_NUMBER,
                        "EXT",
                        Columns.COL_ALLOY,
                        Columns.COL_CAST_TYPE,
                        "Quantity of Molds",
                        "Castings Per Mold",
                        "Quantity of Cores",
                        "Total Weight per EXT",
                        "Molds for EXT",
                        "Heat #",
                        "Pour Schedule Day",
                        "Pour Date",
                        "Pour Weekday",
                        "Pour Buffer Days",
                        "Due Buffer Status",
                        "Planner Diagnostic",
                    ]
                ].copy(),
                "weight_total": weight_total,
                "mold_total": mold_total,
            }

        return export_blocks
    except Exception as exc:
        raise RuntimeError("Failed while building export blocks") from exc


def print_export_blocks(export_blocks):
    """Print a formatted console preview of all export blocks."""
    try:
        for day in export_blocks:
            print("\n" + "=" * 50)
            print(
                f"Mold Schedule    "
                f"{export_blocks[day]['date'].strftime('%m/%d/%Y')}    "
                f"{export_blocks[day]['weekday']}"
            )
            print("=" * 50)
            print(export_blocks[day]["rows"])
            print(f"\nTOTAL WEIGHT: {export_blocks[day]['weight_total']}")
            print(f"TOTAL MOLDS: {export_blocks[day]['mold_total']}")
    except Exception as exc:
        raise RuntimeError("Failed while printing export blocks") from exc


def _resolve_column_value(row, *candidates):
    """Return the first populated value from a list of accepted field names."""
    for candidate in candidates:
        if candidate in row:
            value = row.get(candidate)
            if value is not None and not (isinstance(value, str) and value.strip() == ""):
                return value
    for candidate in candidates:
        if candidate in row:
            return row.get(candidate)
    return ""


def build_excel_rows(export_blocks):
    """
    Flatten export_blocks into a list of row lists for simple sequential writing.

    Each day contributes: a header row, a column-label row, data rows, a TOTALS
    row, and one blank spacer row.

    This is an alternate representation used for verification and lightweight
    consumers; Export_Mold_Schedule writes directly from structured blocks.

    Returns:
        list of lists – each inner list represents one Excel row.
    """
    excel_rows = []

    for day in sorted(export_blocks.keys()):
        block = export_blocks[day]
        excel_rows.append(["Mold Schedule", block["date"].strftime("%m/%d/%Y"), block["weekday"]])
        excel_rows.append([
            "Customer Name",
            "Part Number",
            "Job Number",
            "EXT",
            "Alloy",
            "Mold Type",
            "Quantity of Molds",
            "Cores Per Mold",
            "Total Weight",
            "Number of Molds Today",
        ])

        for _, row in block["rows"].iterrows():
            excel_rows.append([
                row.get("Customer Name", ""),
                row.get("Part Number", ""),
                row.get(Columns.COL_JOB_NUMBER, ""),
                row.get("EXT", ""),
                row.get(Columns.COL_ALLOY, ""),
                row.get(Columns.COL_CAST_TYPE, ""),
                row.get("Quantity of Molds", ""),
                row.get("Quantity of Cores", ""),
                _resolve_column_value(row, "Total Weight", "Total Weight per EXT"),
                _resolve_column_value(row, "Number of Molds Today", "# of Molds Today", "Molds for EXT"),
            ])

        excel_rows.append(["TOTALS", "", "", "", "", "", "", "", block["weight_total"], block["mold_total"]])
        excel_rows.append([])

    return excel_rows


def _build_simple_melt_dept_rows(melt_schedule, day_dates):
    """Return simplified pour-planning rows for the melt department sheet."""
    if not melt_schedule:
        return []

    melt_dept_rows = []
    for day in sorted(melt_schedule.keys()):
        day_rows = melt_schedule[day].get("rows", pd.DataFrame()).copy()
        if day_rows.empty:
            continue

        for _, planned_row in day_rows.iterrows():
            job_number = str(planned_row.get(Columns.COL_JOB_NUMBER, "") or "").strip()
            alloy = str(planned_row.get(Columns.COL_ALLOY, "") or "").strip()
            customer_name = str(planned_row.get("Customer Name", "") or "").strip()
            heat_number = planned_row.get("Heat #", "")
            row_molds = float(pd.to_numeric(planned_row.get("Molds for EXT", 0), errors="coerce") or 0)
            row_weight = float(pd.to_numeric(planned_row.get("Total Weight per EXT", 0), errors="coerce") or 0)
            melt_dept_rows.append(
                {
                    "Pour Date": _normalize_due_date(day_dates.get(day, {}).get("date", pd.NaT)),
                    "Weekday": day_dates.get(day, {}).get("weekday", ""),
                    "Heat #": heat_number,
                    "Alloy": alloy,
                    "Job Number": job_number,
                    "Molds on Floor": row_molds,
                    "Pour Weight Required (lbs)": row_weight,
                    "Customer Name": customer_name,
                }
            )

    if not melt_dept_rows:
        return []

    grouped = pd.DataFrame(melt_dept_rows)
    grouped["Molds on Floor"] = pd.to_numeric(grouped["Molds on Floor"], errors="coerce").fillna(0)
    grouped["Pour Weight Required (lbs)"] = pd.to_numeric(grouped["Pour Weight Required (lbs)"], errors="coerce").fillna(0)
    grouped = grouped.groupby(
        ["Pour Date", "Weekday", "Heat #", "Alloy", "Job Number", "Customer Name"],
        dropna=False,
        as_index=False,
    ).agg(
        **{
            "Molds on Floor": ("Molds on Floor", "sum"),
            "Pour Weight Required (lbs)": ("Pour Weight Required (lbs)", "sum"),
        }
    )
    return grouped.to_dict(orient="records")


def _write_day_blocked_melt_sheet(ws, melt_schedule, day_dates, header_title, created_on):
    """Write melt rows grouped by pour day in the same day-block style as the mold schedule."""
    current_row = 5
    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    bold = Font(bold=True)
    day_title_fill = REPORT_HEADER_FILL

    _write_schedule_banner(
        ws,
        header_title,
        created_on,
        "Average Poured Lbs / Day",
        "0.0",
        10,
    )

    melt_headers = [
        "Pour Date",
        "Weekday",
        "Heat #",
        "Alloy",
        "Job Number",
        "Molds on Floor",
        "Pour Weight Required (lbs)",
        "Customer Name",
    ]

    for day in sorted(melt_schedule.keys()):
        day_rows = melt_schedule[day].get("rows", pd.DataFrame()).copy()
        if day_rows.empty:
            continue

        ws.cell(current_row, 1, "Melt Schedule")
        ws.cell(current_row, 2, day_dates.get(day, {}).get("date", "").strftime("%m/%d/%Y") if day_dates.get(day, {}).get("date") is not None else "")
        ws.cell(current_row, 4, day_dates.get(day, {}).get("weekday", ""))

        for col in range(1, 9):
            ws.cell(current_row, col).font = bold
            ws.cell(current_row, col).fill = day_title_fill
        ws.row_dimensions[current_row].height = 20
        current_row += 2

        for col_num, header in enumerate(melt_headers, start=1):
            cell = ws.cell(current_row, col_num, header)
            cell.border = thin
            cell.font = bold
        _apply_schedule_day_header(ws, current_row, len(melt_headers))
        current_row += 1

        for _, planned_row in day_rows.iterrows():
            molds_value = planned_row.get("Molds for EXT", 0)
            weight_value = planned_row.get("Total Weight per EXT", 0)
            molds_numeric = pd.to_numeric(molds_value, errors="coerce")
            if pd.isna(molds_numeric):
                molds_numeric = 0
            weight_numeric = pd.to_numeric(weight_value, errors="coerce")
            if pd.isna(weight_numeric):
                weight_numeric = 0

            values = [
                _normalize_due_date(day_dates.get(day, {}).get("date", pd.NaT)),
                day_dates.get(day, {}).get("weekday", ""),
                planned_row.get("Heat #", ""),
                str(planned_row.get(Columns.COL_ALLOY, "") or "").strip(),
                str(planned_row.get(Columns.COL_JOB_NUMBER, "") or "").strip(),
                molds_numeric,
                weight_numeric,
                str(planned_row.get("Customer Name", "") or "").strip(),
            ]
            for col_num, value in enumerate(values, start=1):
                cell = ws.cell(current_row, col_num, value)
                cell.border = thin
                cell.fill = REPORT_SEPARATOR_FILL
                if col_num == 1 and value != "":
                    cell.number_format = "m/d/yyyy"
                if col_num in {6, 7} and value != "":
                    cell.number_format = "#,##0.00" if col_num == 7 else "#,##0"
                if col_num in {6, 7}:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

        current_row += 2

    widths = {
        "A": 11,
        "B": 11,
        "C": 8,
        "D": 11,
        "E": 12,
        "F": 14,
        "G": 20,
        "H": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    _apply_11x17_portrait_layout(ws)


def export_mold_schedule(Export_Blocks, output_file="Mold Schedule.xlsx", job_shipping_rows=None, melt_schedule=None, day_dates=None):
    """
    Write the mold schedule to an Excel workbook.

    Each production day occupies its own block of rows separated by two blank
    rows.  Column widths, bold headers, and thin borders are applied.

    Args:
        Export_Blocks: Output of Build_Daily_Export_Blocks.
        output_file:   Destination path for the workbook.
        melt_schedule: Optional melt schedule used to add a simplified Melt Schedule sheet.
        day_dates:     Optional day-date mapping for melt schedule rows.
    """
    try:
        _ensure_output_parent(output_file)
        wb = Workbook()
        ws = wb.active
        ws.title = "Mold Schedule"
        created_on = datetime.now()
        total_molds = sum(float(block.get("mold_total", 0) or 0) for block in Export_Blocks.values())
        avg_molds_per_day = total_molds / max(len(Export_Blocks), 1)

        _write_schedule_banner(
            ws,
            "Mold Schedule",
            created_on,
            "Average Molds / Day",
            f"{avg_molds_per_day:,.1f}",
            10,
        )

        current_row = 5

        thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        bold = Font(bold=True)
        day_title_fill = REPORT_HEADER_FILL

        for day in sorted(Export_Blocks.keys()):
            block = Export_Blocks[day]

            ws.cell(current_row, 1, "Mold Schedule")
            ws.cell(current_row, 2, block["date"].strftime("%m/%d/%Y"))
            ws.cell(current_row, 4, block["weekday"])

            for col in range(1, 11):
                ws.cell(current_row, col).font = bold
                ws.cell(current_row, col).fill = day_title_fill

            ws.row_dimensions[current_row].height = 20

            current_row += 2

            headers = [
                "Customer Name",
                "Part Number",
                "Job Number",
                "EXT",
                "Alloy",
                "Mold Type",
                "Quantity of Molds",
                "Quantity of Cores",
                "Total Weight",
                "Number of Molds Today",
            ]

            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(current_row, col_num, header)
                cell.border = thin
                cell.font = bold

            _apply_schedule_day_header(ws, current_row, len(headers))

            current_row += 1

            for _, row in block["rows"].iterrows():
                values = [
                    row.get("Customer Name", ""),
                    row.get("Part Number", ""),
                    row.get(Columns.COL_JOB_NUMBER, ""),
                    row.get("EXT", ""),
                    row.get(Columns.COL_ALLOY, ""),
                    row.get(Columns.COL_CAST_TYPE, ""),
                    row.get("Quantity of Molds", ""),
                    row.get("Quantity of Cores", ""),
                    _resolve_column_value(row, "Total Weight", "Total Weight per EXT"),
                    _resolve_column_value(row, "Number of Molds Today", "# of Molds Today", "Molds for EXT"),
                ]

                for col_num, value in enumerate(values, start=1):
                    cell = ws.cell(current_row, col_num, value)
                    cell.border = thin
                    cell.fill = REPORT_SEPARATOR_FILL

                current_row += 1

            ws.cell(current_row, 1, "TOTALS")
            ws.cell(current_row, 9, block["weight_total"])
            ws.cell(current_row, 10, block["mold_total"])
            _apply_schedule_total_row(ws, current_row, [1, 9, 10])
            current_row += 4

        widths = {
            "A": 20,
            "B": 13,
            "C": 11,
            "D": 11,
            "E": 10,
            "F": 5,
            "G": 10,
            "H": 10,
            "I": 13,
            "J": 11,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        _apply_11x17_portrait_layout(ws)

        if melt_schedule and day_dates:
            ws_melt = wb.create_sheet("Melt Schedule")
            _write_day_blocked_melt_sheet(ws_melt, melt_schedule, day_dates, "Melt Schedule", created_on)

        if job_shipping_rows:
            ws_jobs = wb.create_sheet("Job Shipping Outlook")
            fill_green = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
            fill_yellow = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
            fill_orange = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
            fill_red = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
            fill_gray = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")

            headers = [
                "Job Number",
                "Schedule Status",
                "Planned Molds",
                "Scheduled Molds",
                "Mold Day",
                "Mold Date",
                "Pour Day",
                "Pour Date",
                "Expected Ship Date",
                "Due Date",
                "Ship Buffer Days",
                "On-Time",
            ]

            for col_num, header in enumerate(headers, start=1):
                cell = ws_jobs.cell(1, col_num, header)
                cell.font = bold
                cell.border = thin

            row_num = 2
            for row in job_shipping_rows:
                values = [
                    row.get("Job Number", ""),
                    row.get("Schedule Status", ""),
                    row.get("Planned Molds", ""),
                    row.get("Scheduled Molds", ""),
                    row.get("Mold Day", ""),
                    row.get("Mold Date", ""),
                    row.get("Pour Day", ""),
                    row.get("Pour Date", ""),
                    row.get("Expected Ship Date", ""),
                    row.get("Due Date", ""),
                    row.get("Ship Buffer Days", ""),
                    row.get("On-Time", ""),
                ]

                for col_num, value in enumerate(values, start=1):
                    cell = ws_jobs.cell(row_num, col_num, value)
                    cell.border = thin
                    if col_num in {6, 8, 9, 10} and value != "":
                        cell.number_format = "m/d/yyyy"

                schedule_status = str(row.get("Schedule Status", "")).strip().upper()
                on_time = str(row.get("On-Time", "")).strip().upper()
                buffer_value = row.get("Ship Buffer Days", None)

                if schedule_status == "SCHEDULED":
                    ws_jobs.cell(row_num, 2).fill = fill_green
                elif schedule_status == "PARTIALLY SCHEDULED":
                    ws_jobs.cell(row_num, 2).fill = fill_yellow
                elif schedule_status == "NOT YET SCHEDULED":
                    ws_jobs.cell(row_num, 2).fill = fill_orange

                if on_time == "YES":
                    ws_jobs.cell(row_num, 12).fill = fill_green
                elif on_time == "NO":
                    ws_jobs.cell(row_num, 12).fill = fill_red
                elif on_time == "NOT SCHEDULED":
                    ws_jobs.cell(row_num, 12).fill = fill_orange
                else:
                    ws_jobs.cell(row_num, 12).fill = fill_gray

                try:
                    if buffer_value is None or pd.isna(buffer_value):
                        ws_jobs.cell(row_num, 11).fill = fill_gray
                    else:
                        buffer_days = int(buffer_value)
                        if buffer_days < 0:
                            ws_jobs.cell(row_num, 11).fill = fill_red
                        elif buffer_days < 4:
                            ws_jobs.cell(row_num, 11).fill = fill_yellow
                        else:
                            ws_jobs.cell(row_num, 11).fill = fill_green
                except (TypeError, ValueError):
                    ws_jobs.cell(row_num, 11).fill = fill_gray

                row_num += 1

            widths = {
                "A": 12,
                "B": 18,
                "C": 12,
                "D": 13,
                "E": 9,
                "F": 11,
                "G": 9,
                "H": 11,
                "I": 14,
                "J": 11,
                "K": 14,
                "L": 10,
            }
            for col, width in widths.items():
                ws_jobs.column_dimensions[col].width = width

            _apply_11x17_portrait_layout(ws_jobs)

            ws_job_status = wb.create_sheet("Overall Job Status")
            job_status_headers = [
                "Job Number",
                "Customer Name",
                "Expected Ship Date",
                "Due Date",
                "Ship Buffer Days",
                "On-Time",
                "Schedule Status",
            ]

            for col_num, header in enumerate(job_status_headers, start=1):
                cell = ws_job_status.cell(1, col_num, header)
                cell.font = bold
                cell.border = thin

            row_num = 2
            for row in job_shipping_rows:
                values = [
                    row.get("Job Number", ""),
                    row.get("Customer Name", ""),
                    row.get("Expected Ship Date", ""),
                    row.get("Due Date", ""),
                    row.get("Ship Buffer Days", ""),
                    row.get("On-Time", ""),
                    row.get("Schedule Status", ""),
                ]

                for col_num, value in enumerate(values, start=1):
                    cell = ws_job_status.cell(row_num, col_num, value)
                    cell.border = thin
                    if col_num in {3, 4} and value != "":
                        cell.number_format = "m/d/yyyy"

                on_time = str(row.get("On-Time", "") or "").strip().upper()
                if on_time == "YES":
                    ws_job_status.cell(row_num, 6).fill = fill_green
                elif on_time == "NO":
                    ws_job_status.cell(row_num, 6).fill = fill_red
                elif on_time == "NOT SCHEDULED":
                    ws_job_status.cell(row_num, 6).fill = fill_orange
                else:
                    ws_job_status.cell(row_num, 6).fill = fill_gray

                row_num += 1

            status_widths = {
                "A": 12,
                "B": 24,
                "C": 14,
                "D": 11,
                "E": 14,
                "F": 10,
                "G": 18,
            }
            for col, width in status_widths.items():
                ws_job_status.column_dimensions[col].width = width

            _apply_letter_portrait_layout(ws_job_status)

        wb.save(output_file)
        print(f"Saved: {output_file}")
    except Exception as exc:
        raise RuntimeError(f"Failed while exporting mold schedule to {output_file}") from exc


def build_heat_summary_rows(export_blocks, mold_schedule_frame=None):
    """
    Aggregate melt-plan summary rows into a dated export-friendly shape.

    Returns:
        list of dicts based on melt_schedule[day]['heat_summary'] plus schedule date metadata.
        When mold_schedule_frame is provided, includes planner-facing mold lead metrics.
    """
    summary_rows = []

    melt_schedule = export_blocks
    day_dates = None
    if isinstance(export_blocks, tuple):
        melt_schedule, day_dates = export_blocks

    if day_dates is None:
        raise RuntimeError("build_heat_summary_rows requires melt_schedule and day_dates")

    mold_lead_by_heat = {}
    if mold_schedule_frame is not None and not mold_schedule_frame.empty:
        lead_frame = mold_schedule_frame.copy()
        lead_frame["_MoldLeadDays"] = (
            pd.to_numeric(lead_frame.get("Pour Schedule Day"), errors="coerce")
            - pd.to_numeric(lead_frame.get("Schedule Day"), errors="coerce")
        )
        lead_frame = lead_frame[lead_frame["_MoldLeadDays"].notna()].copy()
        if not lead_frame.empty:
            grouped_lead = (
                lead_frame
                .groupby(["Pour Schedule Day", "Heat #"], dropna=False)
                .agg(
                    MaxMoldLeadDays=("_MoldLeadDays", "max"),
                    AvgMoldLeadDays=("_MoldLeadDays", "mean"),
                )
                .reset_index()
            )
            for _, lead_row in grouped_lead.iterrows():
                mold_lead_by_heat[(
                    int(lead_row["Pour Schedule Day"]),
                    lead_row["Heat #"],
                )] = {
                    "Max Mold Lead Days": int(round(float(lead_row["MaxMoldLeadDays"]))),
                    "Avg Mold Lead Days": round(float(lead_row["AvgMoldLeadDays"]), 1),
                }

    for day in sorted(melt_schedule.keys()):
        heat_summary = melt_schedule[day].get("heat_summary", pd.DataFrame()).copy()
        planned_rows = melt_schedule[day].get("rows", pd.DataFrame()).copy()
        if heat_summary.empty:
            continue

        block_date = day_dates[day]["date"]
        block_weekday = day_dates[day]["weekday"]

        heat_breakout_map = {}
        heat_min_days_until_due = {}
        if not planned_rows.empty and "Heat #" in planned_rows.columns:
            for heat_number, heat_df in planned_rows.groupby("Heat #", sort=True):
                if pd.isna(heat_number) or heat_number == "":
                    continue

                due_days = pd.to_numeric(
                    heat_df.get("Days Until Due", pd.Series(dtype="float64")),
                    errors="coerce",
                ).dropna()
                heat_min_days_until_due[heat_number] = (
                    int(due_days.min()) if not due_days.empty else None
                )

                detail_rows = []
                for _, planned_row in heat_df.iterrows():
                    job_number = str(planned_row.get(Columns.COL_JOB_NUMBER, "") or "").strip()
                    ext = str(planned_row.get("EXT", "") or "").strip()
                    due_date = _normalize_due_date(planned_row.get(Columns.COL_DUE_DATE, ""))
                    due_date_text = due_date.strftime("%m/%d/%Y") if hasattr(due_date, "strftime") else ""
                    molds_for_row = pd.to_numeric(
                        planned_row.get("Molds for EXT", 0),
                        errors="coerce",
                    )
                    molds_text = "" if pd.isna(molds_for_row) else str(int(molds_for_row))

                    job_label = f"{job_number}-{ext}" if job_number and ext else job_number
                    detail_rows.append(
                        f"{job_label} | Due {due_date_text} | Molds {molds_text}"
                    )

                heat_breakout_map[heat_number] = "; ".join(detail_rows)

        for _, row in heat_summary.iterrows():
            heat_number = row.get("Heat #", "")
            schedule_date = block_date.date() if hasattr(block_date, "date") else block_date
            earliest_due = _normalize_date_value(row.get("Earliest Due Date", ""))
            latest_due = _normalize_date_value(row.get("Latest Due Date", ""))
            schedule_ts = _normalize_date_value(schedule_date)
            due_buffer_days = None
            if not pd.isna(earliest_due) and not pd.isna(schedule_ts):
                due_buffer_days = int((earliest_due - schedule_ts).days)
            buffer_status, diagnostic = _planner_risk_for_buffer(due_buffer_days)
            lead_metrics = mold_lead_by_heat.get((int(day), heat_number), {})
            min_days_until_due = heat_min_days_until_due.get(heat_number, None)

            if due_buffer_days is None:
                two_week_status = "UNKNOWN"
                two_week_note = "Cannot evaluate two-week rule without due date"
            elif due_buffer_days >= 14:
                two_week_status = "PASS"
                two_week_note = "Pour is at least 14 days before due"
            elif min_days_until_due is not None and min_days_until_due < 14:
                two_week_status = "EXCEPTION - NO TIME"
                two_week_note = "Due is already inside 14 days; poured at earliest feasible slot"
            else:
                two_week_status = "VIOLATION"
                two_week_note = "Pour is inside 14-day window even though due runway exists"

            summary_rows.append(
                {
                    "Schedule Date": schedule_date,
                    "Weekday": block_weekday,
                    "Heat Slot": row.get("Heat Slot", ""),
                    "Heat #": heat_number,
                    "Heat Status": row.get("Heat Status", ""),
                    "Planning Priority": row.get("Planning Priority", ""),
                    "Review Window": row.get("Review Window", ""),
                    "Anchor Alloy": row.get("Anchor Alloy", ""),
                    "Compatibility Group": row.get("Compatibility Group", ""),
                    "Earliest Due Date": _normalize_due_date(row.get("Earliest Due Date", "")),
                    "Latest Due Date": _normalize_due_date(row.get("Latest Due Date", "")),
                    "Target Pour Date": _normalize_due_date(_target_pour_date(earliest_due)),
                    "Pour Buffer Days": due_buffer_days,
                    "Due Buffer Status": buffer_status,
                    "Total Weight (lbs)": float(row.get("Total Weight (lbs)", 0) or 0),
                    "Total Molds": float(row.get("Total Molds", 0) or 0),
                    "Rows in Heat": int(row.get("Rows in Heat", 0) or 0),
                    "Jobs": row.get("Jobs", ""),
                    "Extensions": row.get("Extensions", ""),
                    "Max Mold Lead Days": lead_metrics.get("Max Mold Lead Days", ""),
                    "Avg Mold Lead Days": lead_metrics.get("Avg Mold Lead Days", ""),
                    "Two Week Rule Status": two_week_status,
                    "Two Week Rule Note": two_week_note,
                    "Job Breakout": heat_breakout_map.get(heat_number, ""),
                    "Planner Diagnostic": diagnostic,
                }
            )

    return summary_rows


def build_heat_planner_rows(summary_rows):
    """Build planner-friendly rows including the blank reserved heat slot."""
    planner_rows = []

    for row in summary_rows:
        planner_rows.append(
            {
                "Schedule Date": row["Schedule Date"],
                "Weekday": row["Weekday"],
                "Heat Slot": row["Heat Slot"],
                "Heat Status": row["Heat Status"],
                "Heat #": row["Heat #"],
                "Planning Priority": row["Planning Priority"],
                "Review Window": row["Review Window"],
                "Anchor Alloy": row["Anchor Alloy"],
                "Compatibility Group": row["Compatibility Group"],
                "Earliest Due Date": row["Earliest Due Date"],
                "Latest Due Date": row["Latest Due Date"],
                "Target Pour Date": row["Target Pour Date"],
                "Pour Buffer Days": row["Pour Buffer Days"],
                "Due Buffer Status": row["Due Buffer Status"],
                "Total Weight (lbs)": row["Total Weight (lbs)"],
                "Total Molds": row["Total Molds"],
                "Jobs": row["Jobs"],
                "Extensions": row["Extensions"],
                "Max Mold Lead Days": row.get("Max Mold Lead Days", ""),
                "Avg Mold Lead Days": row.get("Avg Mold Lead Days", ""),
                "Two Week Rule Status": row.get("Two Week Rule Status", ""),
                "Two Week Rule Note": row.get("Two Week Rule Note", ""),
                "Planner Diagnostic": row["Planner Diagnostic"],
                "Manual Alloy": "",
                "Manual Weight (lbs)": "",
                "Manual Molds": "",
                "Planner Notes": "",
            }
        )

    return planner_rows


def build_heat_daily_totals_rows(summary_rows):
    """
    Roll up heat summary rows to one row per production day.

    Returns:
        list of dicts with keys: Schedule Date, Weekday, Total Heats,
        Total Weight (lbs), Total Molds.
    """
    if not summary_rows:
        return []

    summary_df = pd.DataFrame(summary_rows)
    summary_df = summary_df[summary_df["Heat Status"] != "Reserved"].copy()
    if summary_df.empty:
        return []

    summary_df["Heat Status"] = summary_df["Heat Status"].fillna("")
    grouped = (
        summary_df
        .groupby(["Schedule Date", "Weekday"], dropna=False, sort=True)
        .agg(
            TotalHeats=("Heat #", "nunique"),
            PlannedHeats=("Heat Status", lambda values: int((values == "Planned").sum())),
            OverflowHeats=("Heat Status", lambda values: int((values == "Overflow").sum())),
            TotalWeightLbs=("Total Weight (lbs)", "sum"),
            TotalMolds=("Total Molds", "sum"),
            AtRiskHeats=("Due Buffer Status", lambda values: int((values == "AT RISK").sum())),
        )
        .reset_index()
    )

    daily_rows = []
    for _, row in grouped.iterrows():
        daily_rows.append(
            {
                "Schedule Date": row["Schedule Date"],
                "Weekday": row["Weekday"],
                "Total Heats": int(row["TotalHeats"]),
                "Planned Heats": int(row["PlannedHeats"]),
                "Overflow Heats": int(row["OverflowHeats"]),
                "Total Weight (lbs)": float(row["TotalWeightLbs"]),
                "Total Molds": float(row["TotalMolds"]),
                "At-Risk Heats": int(row["AtRiskHeats"]),
            }
        )

    return daily_rows


def build_heat_detail_rows(melt_schedule, day_dates):
    """Flatten per-row planned heat assignments for workbook export."""
    detail_rows = []

    for day in sorted(melt_schedule.keys()):
        day_plan = melt_schedule.get(day, {})
        planned_rows = day_plan.get("rows", pd.DataFrame()).copy()
        if planned_rows.empty:
            continue

        block_date = day_dates.get(day, {}).get("date", pd.NaT)
        schedule_date = block_date.date() if hasattr(block_date, "date") else block_date
        schedule_weekday = day_dates.get(day, {}).get("weekday", "")

        for _, row in planned_rows.iterrows():
            due_ts = _normalize_date_value(row.get(Columns.COL_DUE_DATE, pd.NaT))
            schedule_ts = _normalize_date_value(schedule_date)
            due_buffer_days = None
            if not pd.isna(due_ts) and not pd.isna(schedule_ts):
                due_buffer_days = int((due_ts - schedule_ts).days)

            row_weight = float(pd.to_numeric(row.get("Total Weight per EXT", 0), errors="coerce") or 0)
            row_molds = int(pd.to_numeric(row.get("Molds for EXT", 0), errors="coerce") or 0)

            detail_rows.append(
                {
                    "Schedule Date": schedule_date,
                    "Weekday": schedule_weekday,
                    "Pour Schedule Day": row.get("Pour Schedule Day", day),
                    "Heat #": row.get("Heat #", ""),
                    "Global Heat #": row.get("Global Heat #", ""),
                    "Planning Priority": row.get("Planning Priority", ""),
                    "Review Window": row.get("Review Window", ""),
                    "Days Until Due": row.get("Days Until Due", ""),
                    "Due Date": _normalize_due_date(row.get(Columns.COL_DUE_DATE, "")),
                    "Due Buffer Days": due_buffer_days,
                    "Compatibility Group": row.get("Compatibility Group", ""),
                    "Alloy": row.get(Columns.COL_ALLOY, ""),
                    "Job Number": row.get(Columns.COL_JOB_NUMBER, ""),
                    "EXT": row.get("EXT", ""),
                    "Extension_Seq": row.get("Extension_Seq", ""),
                    "Molds for EXT": row_molds,
                    "Total Weight per EXT": row_weight,
                }
            )

    return detail_rows


def _format_date_range(values):
    """Return m/d/yyyy or m/d/yyyy - m/d/yyyy for a list of date-like values."""
    normalized = []
    for value in values:
        parsed = _normalize_date_value(value)
        if not pd.isna(parsed):
            normalized.append(parsed.date())

    if not normalized:
        return ""

    first_date = min(normalized)
    last_date = max(normalized)
    if first_date == last_date:
        return first_date.strftime("%m/%d/%Y")
    return f"{first_date:%m/%d/%Y} - {last_date:%m/%d/%Y}"


def export_heat_summary(
    melt_schedule,
    day_dates,
    output_file="Heat Summary.xlsx",
    mold_schedule_frame=None,
    mold_day_dates=None,
):
    """
    Write the heat summary workbook.

    Sheet 1 "Heat Summary"       – day-block melt schedule layout.
    Sheet 2 "Heat Planner"       – planner-facing worksheet with manual fill columns.
    Sheet 3 "Melt Mgmt Summary"  – management view of planned pours.
    Sheet 4 "Melt Dept Schedule" - melt-department execution schedule.
    Sheet 5 "Melt Diagnostics"   - skipped pour-day diagnostics and likely causes.

    Args:
        melt_schedule: Output of build_melt_schedule.
        day_dates:     Output of Build_Schedule_Dates.
        output_file:   Destination path for the workbook.
        mold_schedule_frame: Backfilled mold assignments aligned to planned heats.
        mold_day_dates: Day-to-date mapping for mold schedule days.
    """
    try:
        _ensure_output_parent(output_file)
        wb = Workbook()
        ws = wb.active
        ws.title = "Heat Summary"

        headers = [
            "Schedule Date",
            "Weekday",
            "Heat Slot",
            "Heat #",
            "Heat Status",
            "Planning Priority",
            "Review Window",
            "Anchor Alloy",
            "Compatibility Group",
            "Earliest Due Date",
            "Latest Due Date",
            "Target Pour Date",
            "Pour Buffer Days",
            "Due Buffer Status",
            "Total Weight (lbs)",
            "Total Molds",
            "Rows in Heat",
            "Jobs",
            "Extensions",
            "Max Mold Lead Days",
            "Avg Mold Lead Days",
            "Two Week Rule Status",
            "Two Week Rule Note",
            "Job Breakout",
            "Planner Diagnostic",
        ]

        bold = Font(bold=True)
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        fill_green = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        fill_yellow = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
        fill_red = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        fill_gray = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")

        def _two_week_status_fill(status):
            status_text = str(status or "").strip().upper()
            if status_text == "PASS":
                return fill_green
            if status_text == "EXCEPTION - NO TIME":
                return fill_yellow
            if status_text == "VIOLATION":
                return fill_red
            return fill_gray

        summary_rows = build_heat_summary_rows((melt_schedule, day_dates), mold_schedule_frame=mold_schedule_frame)
        total_poured_lbs = 0.0
        day_count = 0
        summary_df = pd.DataFrame(summary_rows)
        if not summary_df.empty:
            grouped_for_totals = summary_df[summary_df["Heat Status"] != "Reserved"].groupby(["Schedule Date", "Weekday"], sort=True)
            day_count = len(grouped_for_totals)
            total_poured_lbs = float(summary_df[summary_df["Heat Status"] != "Reserved"]["Total Weight (lbs)"].sum())

        avg_poured_lbs_per_day = total_poured_lbs / max(day_count, 1)

        _write_schedule_banner(
            ws,
            "Heat Schedule",
            datetime.now(),
            "Average Poured Lbs / Day",
            f"{avg_poured_lbs_per_day:,.1f}",
            len(headers),
        )

        current_row = 5

        if not summary_df.empty:
            grouped = summary_df.groupby(["Schedule Date", "Weekday"], sort=True)
            for (schedule_date, weekday), day_df in grouped:
                ws.cell(current_row, 1, "Heat Schedule")
                ws.cell(current_row, 2, schedule_date)
                ws.cell(current_row, 4, weekday)
                ws.cell(current_row, 2).number_format = "m/d/yyyy"
                for col in range(1, len(headers) + 1):
                    ws.cell(current_row, col).font = bold
                    ws.cell(current_row, col).fill = REPORT_HEADER_FILL

                current_row += 2

                for col_num, header in enumerate(headers, start=1):
                    cell = ws.cell(current_row, col_num, header)
                    cell.font = bold
                    cell.border = thin

                _apply_schedule_day_header(ws, current_row, len(headers))

                current_row += 1

                for _, row in day_df.iterrows():
                    values = [
                        row["Schedule Date"],
                        row["Weekday"],
                        row["Heat Slot"],
                        row["Heat #"],
                        row["Heat Status"],
                        row["Planning Priority"],
                        row["Review Window"],
                        row["Anchor Alloy"],
                        row["Compatibility Group"],
                        row["Earliest Due Date"],
                        row["Latest Due Date"],
                        row["Target Pour Date"],
                        row["Pour Buffer Days"],
                        row["Due Buffer Status"],
                        row["Total Weight (lbs)"],
                        row["Total Molds"],
                        row["Rows in Heat"],
                        row["Jobs"],
                        row["Extensions"],
                        row.get("Max Mold Lead Days", ""),
                        row.get("Avg Mold Lead Days", ""),
                        row.get("Two Week Rule Status", ""),
                        row.get("Two Week Rule Note", ""),
                        row["Job Breakout"],
                        row["Planner Diagnostic"],
                    ]

                    for col_num, value in enumerate(values, start=1):
                        cell = ws.cell(current_row, col_num, value)
                        cell.border = thin
                        cell.fill = REPORT_SEPARATOR_FILL
                        if col_num in {1, 10, 11, 12} and value != "":
                            cell.number_format = "m/d/yyyy"
                        if col_num == 15:
                            cell.number_format = "#,##0.00"
                        if col_num == 16:
                            cell.number_format = "#,##0"
                        if col_num in {23, 24, 25}:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")

                    ws.cell(current_row, 22).fill = _two_week_status_fill(row.get("Two Week Rule Status", ""))

                    current_row += 1

                ws.cell(current_row, 1, "TOTALS")
                ws.cell(current_row, 15, float(day_df["Total Weight (lbs)"].sum()))
                ws.cell(current_row, 16, float(day_df["Total Molds"].sum()))
                _apply_schedule_total_row(ws, current_row, [1, 15, 16])
                current_row += 4

        widths = {
            "A": 11,
            "B": 11,
            "C": 8,
            "D": 7,
            "E": 11,
            "F": 13,
            "G": 12,
            "H": 10,
            "I": 12,
            "J": 10,
            "K": 10,
            "L": 10,
            "M": 9,
            "N": 11,
            "O": 12,
            "P": 10,
            "Q": 9,
            "R": 16,
            "S": 18,
            "T": 14,
            "U": 14,
            "V": 16,
            "W": 36,
            "X": 30,
            "Y": 26,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        _apply_11x17_portrait_layout(ws)

        ws_planner = wb.create_sheet("Heat Planner")
        planner_headers = [
            "Schedule Date",
            "Weekday",
            "Heat Slot",
            "Heat Status",
            "Heat #",
            "Planning Priority",
            "Review Window",
            "Anchor Alloy",
            "Compatibility Group",
            "Earliest Due Date",
            "Latest Due Date",
            "Target Pour Date",
            "Pour Buffer Days",
            "Due Buffer Status",
            "Total Weight (lbs)",
            "Total Molds",
            "Jobs",
            "Extensions",
            "Max Mold Lead Days",
            "Avg Mold Lead Days",
            "Two Week Rule Status",
            "Two Week Rule Note",
            "Planner Diagnostic",
            "Manual Alloy",
            "Manual Weight (lbs)",
            "Manual Molds",
            "Planner Notes",
        ]

        for col_num, header in enumerate(planner_headers, start=1):
            cell = ws_planner.cell(1, col_num, header)
            cell.font = bold
            cell.border = thin

        planner_rows = build_heat_planner_rows(summary_rows)
        current_row = 2

        for row in planner_rows:
            values = [
                row["Schedule Date"],
                row["Weekday"],
                row["Heat Slot"],
                row["Heat Status"],
                row["Heat #"],
                row["Planning Priority"],
                row["Review Window"],
                row["Anchor Alloy"],
                row["Compatibility Group"],
                row["Earliest Due Date"],
                row["Latest Due Date"],
                row["Target Pour Date"],
                row["Pour Buffer Days"],
                row["Due Buffer Status"],
                row["Total Weight (lbs)"],
                row["Total Molds"],
                row["Jobs"],
                row["Extensions"],
                row.get("Max Mold Lead Days", ""),
                row.get("Avg Mold Lead Days", ""),
                row.get("Two Week Rule Status", ""),
                row.get("Two Week Rule Note", ""),
                row["Planner Diagnostic"],
                row["Manual Alloy"],
                row["Manual Weight (lbs)"],
                row["Manual Molds"],
                row["Planner Notes"],
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws_planner.cell(current_row, col_num, value)
                cell.border = thin
                if col_num in {1, 10, 11, 12} and value != "":
                    cell.number_format = "m/d/yyyy"
                if col_num in {15, 25} and value != "":
                    cell.number_format = "#,##0.00"
                if col_num in {16, 26} and value != "":
                    cell.number_format = "#,##0"
                if col_num in {22, 23, 27}:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            ws_planner.cell(current_row, 21).fill = _two_week_status_fill(row.get("Two Week Rule Status", ""))

            current_row += 1

        planner_widths = {
            "A": 14,
            "B": 12,
            "C": 10,
            "D": 12,
            "E": 8,
            "F": 18,
            "G": 15,
            "H": 15,
            "I": 18,
            "J": 14,
            "K": 14,
            "L": 14,
            "M": 12,
            "N": 14,
            "O": 18,
            "P": 12,
            "Q": 24,
            "R": 28,
            "S": 14,
            "T": 14,
            "U": 18,
            "V": 36,
            "W": 30,
            "X": 15,
            "Y": 18,
            "Z": 14,
            "AA": 28,
        }

        for col, width in planner_widths.items():
            ws_planner.column_dimensions[col].width = width

        _apply_11x17_portrait_layout(ws_planner)

        ws_melt_mgmt = wb.create_sheet("Melt Mgmt Summary")
        melt_mgmt_headers = [
            "Pour Date",
            "Heat #",
            "Customer(s)",
            "Poured Job IDs",
            "Due Date(s)",
            "Mold Date(s)",
            "Pour Date(s)",
            "Expected Ship Date",
            "Pour Buffer Days",
            "Due Buffer Status",
            "Two Week Rule Status",
            "Planner Diagnostic",
        ]
        for col_num, header in enumerate(melt_mgmt_headers, start=1):
            cell = ws_melt_mgmt.cell(1, col_num, header)
            cell.font = bold
            cell.border = thin

        customer_map = {}
        poured_jobs_map = {}
        for day in sorted(melt_schedule.keys()):
            day_rows = melt_schedule[day].get("rows", pd.DataFrame()).copy()
            if day_rows.empty:
                continue
            for heat_number, heat_df in day_rows.groupby("Heat #", sort=True):
                if pd.isna(heat_number) or heat_number == "":
                    continue
                names = [
                    str(value).strip()
                    for value in heat_df.get("Customer Name", pd.Series(dtype="object"))
                    if str(value).strip()
                ]
                unique_names = []
                for name in names:
                    if name not in unique_names:
                        unique_names.append(name)
                customer_map[(int(day), heat_number)] = ", ".join(unique_names)

                jobs = [
                    str(value).strip()
                    for value in heat_df.get(Columns.COL_JOB_NUMBER, pd.Series(dtype="object"))
                    if str(value).strip()
                ]
                unique_jobs = []
                for job in jobs:
                    if job not in unique_jobs:
                        unique_jobs.append(job)
                poured_jobs_map[(int(day), heat_number)] = ", ".join(unique_jobs)

        mold_date_map = {}
        if mold_schedule_frame is not None and not mold_schedule_frame.empty:
            mold_frame = mold_schedule_frame.copy()
            for (pour_day, heat_number), mold_df in mold_frame.groupby(["Pour Schedule Day", "Heat #"], sort=True):
                day_values = pd.to_numeric(mold_df.get("Schedule Day", pd.Series(dtype="float64")), errors="coerce").dropna()
                mold_dates = []
                for day_value in day_values:
                    day_int = int(day_value)
                    if mold_day_dates and day_int in mold_day_dates:
                        mold_dates.append(mold_day_dates[day_int].get("date", pd.NaT))
                if mold_dates:
                    mold_date_map[(int(pour_day), heat_number)] = _format_date_range(mold_dates)

        fill_green = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
        fill_yellow = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
        fill_red = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")
        fill_gray = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")

        def _status_fill_for_due_buffer(status):
            text = str(status or "").strip().upper()
            if text == "ON TRACK":
                return fill_green
            if text == "WATCH":
                return fill_yellow
            if text == "AT RISK":
                return fill_red
            return fill_gray

        def _status_fill_for_two_week(status):
            text = str(status or "").strip().upper()
            if text == "PASS":
                return fill_green
            if text == "EXCEPTION - NO TIME":
                return fill_yellow
            if text == "VIOLATION":
                return fill_red
            return fill_gray

        mgmt_row = 2
        for row in summary_rows:
            if str(row.get("Heat Status", "")) == "Reserved":
                continue

            pour_date = _normalize_due_date(row.get("Schedule Date", ""))
            due_dates = _format_date_range([
                row.get("Earliest Due Date", ""),
                row.get("Latest Due Date", ""),
            ])
            pour_dates = _format_date_range([row.get("Schedule Date", "")])
            expected_ship = ""
            schedule_ts = _normalize_date_value(row.get("Schedule Date", ""))
            if not pd.isna(schedule_ts):
                expected_ship = (schedule_ts + pd.Timedelta(days=14)).date()

            # Resolve by schedule day and heat number, matching build_heat_summary_rows keying.
            by_day_key = None
            for day in sorted(melt_schedule.keys()):
                day_date = day_dates.get(day, {}).get("date", pd.NaT)
                day_date_norm = _normalize_date_value(day_date)
                row_date_norm = _normalize_date_value(row.get("Schedule Date", pd.NaT))
                if not pd.isna(day_date_norm) and not pd.isna(row_date_norm) and day_date_norm == row_date_norm:
                    by_day_key = (int(day), row.get("Heat #", ""))
                    break

            values = [
                pour_date,
                row.get("Heat #", ""),
                customer_map.get(by_day_key, ""),
                poured_jobs_map.get(by_day_key, ""),
                due_dates,
                mold_date_map.get(by_day_key, ""),
                pour_dates,
                _normalize_due_date(expected_ship),
                row.get("Pour Buffer Days", ""),
                row.get("Due Buffer Status", ""),
                row.get("Two Week Rule Status", ""),
                row.get("Planner Diagnostic", ""),
            ]

            for col_num, value in enumerate(values, start=1):
                cell = ws_melt_mgmt.cell(mgmt_row, col_num, value)
                cell.border = thin
                if col_num in {1, 8} and value != "":
                    cell.number_format = "m/d/yyyy"
                if col_num == 12:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            ws_melt_mgmt.cell(mgmt_row, 10).fill = _status_fill_for_due_buffer(row.get("Due Buffer Status", ""))
            ws_melt_mgmt.cell(mgmt_row, 11).fill = _status_fill_for_two_week(row.get("Two Week Rule Status", ""))
            mgmt_row += 1

        melt_mgmt_widths = {
            "A": 11,
            "B": 8,
            "C": 26,
            "D": 18,
            "E": 23,
            "F": 20,
            "G": 13,
            "H": 13,
            "I": 12,
            "J": 13,
            "K": 18,
            "L": 36,
        }
        for col, width in melt_mgmt_widths.items():
            ws_melt_mgmt.column_dimensions[col].width = width

        melt_capacity_rows = _build_melt_capacity_rows(melt_schedule)
        _write_capacity_block(
            ws_melt_mgmt,
            max(4, mgmt_row + 2),
            "Daily Melt Capacity",
            ["Day", "Heat Capacity", "Weight Capacity"],
            melt_capacity_rows,
            bold,
            thin,
        )
        _apply_letter_portrait_layout(ws_melt_mgmt)

        ws_melt_dept = wb.create_sheet("Melt Dept Schedule")
        _write_day_blocked_melt_sheet(
            ws_melt_dept,
            melt_schedule,
            day_dates,
            "Melt Schedule",
            datetime.now(),
        )

        ws_melt_diag = wb.create_sheet("Melt Diagnostics")
        melt_diag_headers = [
            "Skipped Pour Day",
            "Skipped Pour Date",
            "Weekday",
            "Original Heats On Day",
            "Heats Pushed Later",
            "Pushed To Pour Days",
            "Mold Window Days",
            "Window Line Usage",
            "Window Floor Usage",
            "Likely Cause",
        ]
        for col_num, header in enumerate(melt_diag_headers, start=1):
            cell = ws_melt_diag.cell(1, col_num, header)
            cell.font = bold
            cell.border = thin

        melt_diag_rows = _build_melt_gap_diagnostics(melt_schedule, day_dates, mold_schedule_frame)
        diag_row = 2
        if melt_diag_rows:
            for row in melt_diag_rows:
                values = [
                    row.get("Skipped Pour Day", ""),
                    row.get("Skipped Pour Date", ""),
                    row.get("Weekday", ""),
                    row.get("Original Heats On Day", ""),
                    row.get("Heats Pushed Later", ""),
                    row.get("Pushed To Pour Days", ""),
                    row.get("Mold Window Days", ""),
                    row.get("Window Line Usage", ""),
                    row.get("Window Floor Usage", ""),
                    row.get("Likely Cause", ""),
                ]
                for col_num, value in enumerate(values, start=1):
                    cell = ws_melt_diag.cell(diag_row, col_num, value)
                    cell.border = thin
                    if col_num == 2 and value != "":
                        cell.number_format = "m/d/yyyy"
                    if col_num == 10:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                diag_row += 1
        else:
            ws_melt_diag.cell(2, 1, "No skipped pour days detected in the current melt schedule range.")
            ws_melt_diag.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(melt_diag_headers))
            ws_melt_diag.cell(2, 1).alignment = Alignment(horizontal="left", vertical="center")

        melt_diag_widths = {
            "A": 15,
            "B": 15,
            "C": 11,
            "D": 20,
            "E": 18,
            "F": 20,
            "G": 15,
            "H": 18,
            "I": 18,
            "J": 70,
        }
        for col, width in melt_diag_widths.items():
            ws_melt_diag.column_dimensions[col].width = width
        _apply_letter_portrait_layout(ws_melt_diag)

        wb.save(output_file)
        print(f"Saved: {output_file}")
    except Exception as exc:
        raise RuntimeError(f"Failed while exporting heat summary to {output_file}") from exc
